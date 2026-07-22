#!/usr/bin/env python3
# SPDX-License-Identifier: MIT
# SPDX-FileCopyrightText: 2026 sean galloway
"""Generate stream_signal_contracts.xlsx for the current STREAM RTL.

Modeled on pumice-ddr2-lpddr2/docs/gen_signal_contracts_kmaps.py.

Builds, from scratch (rerunnable / idempotent):
  * CONTRACT sheets -- per interface group, a table of Signal / Width / Dir /
    Driver / "Legal / Correct behavior": when the signal may assert, what it
    must hold across, what violating it means. Rows that were the subject of a
    known bug fix say so (known_issues reference) -- those are the rows a
    future reader needs most.
  * K-MAP sheets -- for each important COMBINATIONAL decision signal in the
    engines / scheduler / descriptor engine / SRAM controller / core monitor
    gating, a Karnaugh map computed from a PYTHON MIRROR of the exact RTL
    expression (file:line cited next to the map, RTL quoted verbatim). Gray
    order 00 01 11 10; >4 vars = one 4x4 grid per page-variable value; 1-cells
    green, 0-cells grey; an "inspection check" note states what a healthy map
    looks like so a bad future edit is visible on sight. Grid and RTL agree BY
    CONSTRUCTION: the map is computed by evaluating the mirrored expression
    over all input combinations, never by hand.

Every citation is VERIFIED at build time: a registry of (file, line, snippet)
is checked against the RTL before the workbook is written, so a rerun after an
RTL edit that moves or changes a cited expression fails loudly instead of
publishing a stale map.

Rerun after engine / scheduler / descriptor / SRAM-controller RTL changes:
    python3 docs/gen_signal_contracts_kmaps.py
"""
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "stream_signal_contracts.xlsx")
REPO = os.path.abspath(os.path.join(HERE, *([".."] * 5)))

GREEN = PatternFill("solid", fgColor="C6EFCE")
GREY = PatternFill("solid", fgColor="F2F2F2")
TITLE = Font(bold=True, size=12)
HDR = Font(bold=True)
MONO = Font(name="Consolas", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Border(*[Side(style="thin")] * 4)

GRAY2 = [(0, 0), (0, 1), (1, 1), (1, 0)]  # gray-code order of 2 bits

# ---------------------------------------------------------------------------
# Citation registry: (path-relative-to-repo, line-number, snippet-that-must-
# appear-on-that-line). Checked before the workbook is written.
# ---------------------------------------------------------------------------
RD_ENG = "projects/components/dmas/stream/rtl/fub/axi_read_engine.sv"
WR_ENG = "projects/components/dmas/stream/rtl/fub/axi_write_engine.sv"
SCHED = "projects/components/dmas/stream/rtl/fub/scheduler.sv"
DESC_ENG = "projects/components/dmas/stream/rtl/fub/descriptor_engine.sv"
SRAM_CTL = "projects/components/dmas/stream/rtl/fub/sram_controller.sv"
SRAM_UNIT = "projects/components/dmas/stream/rtl/fub/sram_controller_unit.sv"
ALLOC = "projects/components/dmas/stream/rtl/fub/stream_alloc_ctrl.sv"
DRAIN = "projects/components/dmas/stream/rtl/fub/stream_drain_ctrl.sv"
CORE = "projects/components/dmas/stream/rtl/macro/stream_core.sv"
RD_MON = "rtl/amba/monitor/axi4_master_rd_mon.sv"
WR_MON = "rtl/amba/monitor/axi4_master_wr_mon.sv"
MON_BASE = "rtl/amba/monitor/axi_monitor_base.sv"
MON_PKG = "rtl/amba/includes/monitor_common_pkg.sv"

KI_WLAST = ("projects/components/dmas/stream/known_issues/resolved/"
            "axi_write_engine_wlast_drain.md")
KI_DBLCNT = ("projects/components/dmas/stream/known_issues/resolved/"
             "sram_drain_bridge_double_count.md")

CITES = [
    (RD_ENG, 316, "w_space_ok[i] = (SCW'(axi_rd_alloc_space_free[i])"),
    (RD_ENG, 327, "w_arb_request[i] = sched_rd_valid[i] && w_space_ok[i]"),
    (RD_ENG, 413, "m_axi_arvalid = w_arb_grant_valid && sched_rd_valid"),
    (RD_ENG, 429, "w_stale_grant = w_arb_grant & ~sched_rd_valid"),
    (RD_ENG, 430, "w_arb_grant_ack = (w_arb_grant & {NC{(m_axi_arvalid"),
    (RD_ENG, 199, "if (m_axi_arvalid && m_axi_arready && (w_arb_grant_id"),
    (RD_ENG, 204, "if (m_axi_rvalid && m_axi_rready && m_axi_rlast"),
    (RD_ENG, 469, "axi_rd_sram_valid = m_axi_rvalid"),
    (RD_ENG, 472, "m_axi_rready = axi_rd_sram_ready"),
    (RD_ENG, 520, "if (m_axi_rvalid && m_axi_rready && (m_axi_rresp != 2'b00))"),
    (WR_ENG, 407, "w_effective_avail[i] = (axi_wr_drain_data_avail[i]"),
    (WR_ENG, 427, "w_has_data[i] = (SCW'(w_effective_avail[i])"),
    (WR_ENG, 430, "w_final_burst[i] = (sched_wr_beats[i] > 0)"),
    (WR_ENG, 434, "w_data_ok[i] = w_has_data[i] || w_final_burst[i]"),
    (WR_ENG, 452, "w_arb_request[i] = sched_wr_valid[i] && w_data_ok[i]"),
    (WR_ENG, 562, "if (w_arb_grant_valid && !r_aw_valid && sched_wr_valid"),
    (WR_ENG, 576, "m_axi_awvalid = r_aw_valid"),
    (WR_ENG, 704, "axi_wr_sram_drain = m_axi_wvalid && m_axi_wready"),
    (WR_ENG, 720, "m_axi_wvalid = r_w_active"),
    (WR_ENG, 725, "m_axi_wlast = (r_w_beats_remaining == 8'd1)"),
    (WR_ENG, 780, "if (!r_w_active && !w_phase_txn_fifo_empty)"),
    (WR_ENG, 784, "else if (m_axi_wvalid && m_axi_wready && m_axi_wlast"),
    (WR_ENG, 941, "m_axi_bready = 1'b1"),
    (WR_ENG, 951, "if (m_axi_bvalid && m_axi_bready && (m_axi_bresp != 2'b00))"),
    (SCHED, 816, "w_read_complete = (r_read_beats_remaining == 32'h0)"),
    (SCHED, 886, "w_write_issued   = (r_write_beats_remaining == 32'h0)"),
    (SCHED, 891, "w_write_complete = (r_write_beats_to_commit == 32'h0)"),
    (SCHED, 893, "w_transfer_complete = w_read_complete && w_write_issued"),
    (SCHED, 911, "w_rd_prefetch_en = cfg_rd_prefetch_enable && !w_is_ext"),
    (SCHED, 916, "w_rd_peek   = w_rd_prefetch_en && w_state_xfer_data"),
    (SCHED, 926, "w_wr_advance = w_rd_prefetch_en && w_state_xfer_data"),
    (SCHED, 947, "w_sched_rd_completing_this_cycle = sched_rd_done_strobe"),
    (SCHED, 962, "sched_rd_valid = (r_current_state == CH_XFER_DATA)"),
    (SCHED, 988, "sched_wr_valid = (r_current_state == CH_XFER_DATA)"),
    (SCHED, 510, "if (w_wr_advance) begin"),
    (SCHED, 512, "end else if (w_transfer_complete && !r_rd_ahead) begin"),
    (SCHED, 527, "if (r_descriptor.next_descriptor_ptr != 32'h0"),
    (SCHED, 529, "end else if (w_write_complete) begin"),
    (SCHED, 1064, "descriptor_ready = (r_current_state == CH_IDLE)"),
    (SCHED, 1102, "if (sched_wr_done_strobe || sched_wr_commit_strobe)"),
    (SCHED, 1147, "w_timeout_expired = cfg_sched_timeout_enable"),
    (SCHED, 1152, "w_timeout_escalate = (cfg_sched_timeout_limit != 8'd0)"),
    (SCHED, 1157, "w_hard_error = descriptor_error || sched_rd_error"),
    (DESC_ENG, 333, "w_apb_skid_valid_in = apb_valid && !r_channel_reset_active"),
    (DESC_ENG, 449, "w_next_addr_valid = (w_next_addr_extended >= cfg_addr0_base"),
    (DESC_ENG, 454, "w_chain_condition = (w_next_addr != '0) && !w_desc_last"),
    (DESC_ENG, 457, "w_chain_eligible = w_chain_condition"),
    (DESC_ENG, 462, "w_desc_committed = (r_current_state == RD_COMPLETE)"),
    (DESC_ENG, 476, "w_prefetch_allows = (w_desc_fifo_count < w_prefetch_limit)"),
    (DESC_ENG, 479, "w_should_chain = w_chain_eligible && w_desc_committed"),
    (DESC_ENG, 485, "w_pending_push_fire = r_chain_pending && w_prefetch_allows"),
    (DESC_ENG, 642, "r_ready = ((r_current_state == RD_WAIT_DATA)"),
    (DESC_ENG, 920, "ar_valid = ((r_current_state == RD_ISSUE_ADDR)"),
    (DESC_ENG, 1037, "descriptor_valid = w_desc_fifo_rd_valid && !r_descriptor_error"),
    (SRAM_UNIT, 141, "rd_valid           (axi_wr_sram_valid && axi_wr_sram_ready)"),
    (SRAM_UNIT, 184, "wr_valid           (axi_rd_sram_valid && axi_rd_sram_ready)"),
    (SRAM_UNIT, 189, "rd_valid           (axi_wr_drain_req)"),
    (SRAM_UNIT, 307, "axi_wr_drain_data_avail = drain_data_available"),
    (ALLOC, 76, "w_write = wr_valid && wr_ready"),
    (ALLOC, 86, "if (w_write && !r_wr_full) begin"),
    (ALLOC, 141, "space_free = (AW+1)'(D) - w_count"),
    (DRAIN, 101, "if (w_read && !r_rd_empty) begin"),
    (DRAIN, 142, "rd_ready = !r_rd_empty"),
    (DRAIN, 145, "data_available = w_count"),
    (CORE, 73, "parameter int RD_MON_MAX_TRANS = NUM_CHANNELS * AR_MAX_OUTSTANDING + 4"),
    (CORE, 660, "int_cfg_rdeng_mon_enable = cfg_rdeng_mon_enable"),
    (CORE, 715, "int_cfg_rdeng_mon_enable = 1'b0"),
    (RD_MON, 477, "fub_axi_arready = w_core_fub_axi_arready &"),
    (WR_MON, 482, "fub_axi_awready = w_core_fub_axi_awready &"),
    (MON_BASE, 483, "block_ready = (MAX_TRANSACTIONS > BLOCK_MARGIN)"),
    (MON_PKG, 115, "return (max_transactions >= 16) ? 2 : 0"),
]


def verify_citations():
    bad = []
    for path, line, snippet in CITES:
        full = os.path.join(REPO, path)
        try:
            with open(full, "r") as f:
                lines = f.readlines()
        except OSError:
            bad.append(f"{path}: unreadable")
            continue
        if line > len(lines) or snippet not in lines[line - 1]:
            bad.append(f"{path}:{line}: expected {snippet!r}")
    if bad:
        print("CITATION DRIFT -- the RTL moved under the cited lines.",
              file=sys.stderr)
        for b in bad:
            print("  " + b, file=sys.stderr)
        sys.exit(1)


def _glabel(bits):
    return "".join(str(b) for b in bits)


class KmapWriter:
    def __init__(self, ws):
        self.ws = ws
        self.row = 1

    def sheet_intro(self, title, lines):
        ws = self.ws
        ws.cell(self.row, 1, title).font = TITLE
        self.row += 1
        for ln in lines:
            c = ws.cell(self.row, 1, ln)
            c.alignment = WRAP
            self.row += 1
        self.row += 1

    def kmap(self, name, source, expr, varnames, fn, check, values=None):
        """One K-map block. varnames: MSB-first list (2..6). fn(*bits)->0/1
        (or a short string when `values` mapping is wanted). Pages over
        varnames[4:]."""
        ws = self.ws
        ws.cell(self.row, 1, name).font = TITLE
        ws.cell(self.row, 4, source).font = MONO
        self.row += 1
        c = ws.cell(self.row, 1, expr)
        c.font = MONO
        c.alignment = WRAP
        self.row += 1
        n = len(varnames)
        rowv = varnames[0:min(2, n)]                  # grid row variables
        colv = varnames[min(2, n):min(4, n)]          # grid col variables
        pagev = varnames[4:]                          # page variables
        ws.cell(self.row, 1,
                f"rows = {'/'.join(rowv)}   cols = {'/'.join(colv)}"
                + (f"   pages = {'/'.join(pagev)}" if pagev else ""))
        self.row += 1
        ws.cell(self.row, 1, f"CHECK BY INSPECTION: {check}").alignment = WRAP
        ws.cell(self.row, 1).font = Font(italic=True)
        self.row += 2

        rows = GRAY2 if len(rowv) == 2 else ([(0,), (1,)] if len(rowv) == 1
                                             else [()])
        cols = GRAY2 if len(colv) == 2 else ([(0,), (1,)] if len(colv) == 1
                                             else [()])
        pages = [()]
        for _ in pagev:
            pages = [p + (b,) for p in pages for b in (0, 1)]

        for page in pages:
            base = self.row
            if pagev:
                lbl = ", ".join(f"{v}={b}" for v, b in zip(pagev, page))
                ws.cell(base, 1, f"[{lbl}]").font = HDR
                base += 1
            # column headers
            for j, cb in enumerate(cols):
                cc = ws.cell(base, 2 + j, _glabel(cb) or "-")
                cc.font = HDR
                cc.alignment = CENTER
            for i, rb in enumerate(rows):
                rc = ws.cell(base + 1 + i, 1, _glabel(rb) or "-")
                rc.font = HDR
                rc.alignment = CENTER
                for j, cb in enumerate(cols):
                    bits = tuple(rb) + tuple(cb) + tuple(page)
                    v = fn(*bits)
                    cell = ws.cell(base + 1 + i, 2 + j)
                    if values is not None:            # multi-valued map
                        cell.value = values.get(v, str(v))
                        benign = str(v) in ("0", "-", "wait", "hold", "IDLE",
                                            "n/a")
                        cell.fill = GREY if benign else GREEN
                    else:
                        cell.value = int(bool(v))
                        cell.fill = GREEN if v else GREY
                    cell.alignment = CENTER
                    cell.border = THIN
            self.row = base + 1 + len(rows) + 1
        self.row += 1

    def table(self, name, source, headers, rows, note=""):
        ws = self.ws
        ws.cell(self.row, 1, name).font = TITLE
        ws.cell(self.row, 4, source).font = MONO
        self.row += 1
        if note:
            c = ws.cell(self.row, 1, note)
            c.alignment = WRAP
            c.font = Font(italic=True)
            self.row += 1
        for j, h in enumerate(headers):
            c = ws.cell(self.row, 1 + j, h)
            c.font = HDR
            c.border = THIN
        self.row += 1
        for r in rows:
            for j, v in enumerate(r):
                c = ws.cell(self.row, 1 + j, v)
                c.border = THIN
                c.alignment = WRAP
            self.row += 1
        self.row += 2


def new_kmap_sheet(wb, name, widths=(30, 9, 9, 9, 9, 9, 9, 9)):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    for col, w in zip("ABCDEFGH", widths):
        ws.column_dimensions[col].width = w
    return KmapWriter(ws)


# ---------------------------------------------------------------------------
# Contract sheets
# ---------------------------------------------------------------------------
CONTRACT_HDRS = ["Group / Channel", "Signal", "Width", "Dir", "Driver",
                 "Legal / Correct behavior (contract)",
                 "Key invariant - assertable check", "Notes / bug history"]


def contract_sheet(wb, name, title, intro, rows):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    for col, w in zip("ABCDEFGH", (16, 28, 10, 6, 16, 58, 46, 46)):
        ws.column_dimensions[col].width = w
    ws.cell(1, 1, title).font = TITLE
    c = ws.cell(2, 1, intro)
    c.alignment = WRAP
    for j, h in enumerate(CONTRACT_HDRS):
        cc = ws.cell(4, 1 + j, h)
        cc.font = HDR
        cc.border = THIN
    r = 5
    for row in rows:
        for j, v in enumerate(row):
            cc = ws.cell(r, 1 + j, v)
            cc.alignment = WRAP
            cc.border = THIN
        r += 1
    return ws


def build_desc_axi_contract(wb):
    rows = [
        ("AR", "m_axi_desc_arvalid", "1", "out", "descriptor_engine FSM",
         "Asserted only in RD_ISSUE_ADDR (chunk 0) or RD_ISSUE_ADDR2 (EXT "
         "chunk 1) and only while !r_axi_read_active; FSM holds the state "
         "until arready, so arvalid is stable until the handshake. At most "
         "one AR in flight per channel engine.",
         "ar_valid |-> state is RD_ISSUE_ADDR/RD_ISSUE_ADDR2; no second AR "
         "until R (rlast) returns",
         f"{DESC_ENG}:920"),
        ("AR", "m_axi_desc_araddr", "64", "out", "descriptor_engine",
         "Descriptor fetch address popped from the descriptor-address FIFO "
         "(APB kick or chain push). Chunk-1 fetch = chunk-0 addr + 0x20. "
         "Chained addresses were range-validated BEFORE being pushed "
         "(cfg_addr0/1 windows); APB addr 0 is rejected at the APB gate.",
         "araddr in [cfg_addr0_base..limit] or [cfg_addr1_base..limit]; "
         "araddr != 0",
         f"{DESC_ENG}:922 (addr mux), :449 (range check)"),
        ("AR", "m_axi_desc_arlen / arsize / arburst", "8/3/2", "out",
         "descriptor_engine",
         "Constants: arlen=0 (single beat), arsize=3'b110 (64 bytes), "
         "arburst=INCR. A descriptor chunk is always one 512-bit beat "
         "carrying the 256-bit descriptor in [255:0].",
         "arlen==0, arsize==6, arburst==01 always",
         f"{DESC_ENG}:924-926"),
        ("AR", "m_axi_desc_arid", "IW", "out", "descriptor_engine",
         "{0, CHANNEL_ID}: the shared descriptor master carries the channel "
         "in the low ID bits so the R beat routes back to the issuing "
         "channel engine.",
         "ar_id[CHAN_WIDTH-1:0] == CHANNEL_ID (formal property in-module)",
         f"{DESC_ENG}:927"),
        ("AR", "arlock/arcache/arprot/arqos/arregion", "-", "out",
         "descriptor_engine",
         "Constants: 0 / 4'b0010 (normal non-cacheable bufferable) / 0 / 0 "
         "/ 0.", "constant", f"{DESC_ENG}:928-932"),
        ("R", "m_axi_desc_rvalid / rready", "1/1", "in/out",
         "fabric / descriptor_engine",
         "r_ready asserted only in RD_WAIT_DATA or RD_WAIT_DATA2 AND only "
         "when r_id matches this channel (w_our_axi_response). Beats for "
         "other channels are NOT accepted by this engine instance - the "
         "shared-master mux relies on that.",
         "r_ready |-> (state==RD_WAIT_DATA||RD_WAIT_DATA2) && "
         "r_id[CHAN_WIDTH-1:0]==CHANNEL_ID",
         f"{DESC_ENG}:642"),
        ("R", "m_axi_desc_rdata", "256", "in", "fabric",
         "One descriptor chunk per beat. Chunk 0 layout: src[63:0], "
         "dst[127:64], length-in-BEATS[159:128], next_ptr[191:160], "
         "valid[192], gen_irq[193], last[194], desc_type[210:208]. "
         "valid bit 0 => r_descriptor_error latched.",
         "descriptor.valid==1 for every consumed descriptor; "
         "length in beats (NOT bytes)",
         f"{DESC_ENG}:840 (valid check)"),
        ("R", "m_axi_desc_rresp", "2", "in", "fabric",
         "OKAY(00) expected. Any other response takes the FSM to RD_ERROR: "
         "descriptor discarded, r_descriptor_error sticky, chain broken "
         "(w_chain_eligible requires !r_descriptor_error).",
         "rresp!=OKAY |-> no FIFO push, descriptor_valid forced 0",
         f"{DESC_ENG}:633, :743"),
        ("R", "m_axi_desc_rlast", "1", "in", "fabric",
         "Single-beat reads: rlast must be 1 on the (only) data beat.",
         "rvalid && our-id |-> rlast", ""),
        ("Monitor", "(desc AXI monitor)", "-", "-", "axi4_master_rd_mon",
         "The descriptor master is wrapped by an axi4_master_rd_mon inside "
         "scheduler_group_array; stream_core defaults its error/timeout/"
         "compl/threshold/debug cones OFF (perf-only) for FPGA fit.",
         "DESC_MON_ENABLE_* params default 0 except perf",
         f"{CORE}:105-115 (param block)"),
    ]
    contract_sheet(
        wb, "Desc AXI master",
        "STREAM descriptor-fetch AXI master (m_axi_desc_*) signal contracts",
        "Shared 256-bit read-only master; one descriptor_engine per channel "
        "issues single-beat fetches tagged with its CHANNEL_ID. Sources: "
        f"{DESC_ENG}, {CORE} ports 284-305.",
        rows)


def build_data_axi_contract(wb):
    rows = [
        ("Data RD / AR", "m_axi_arvalid", "1", "out", "axi_read_engine",
         "COMBINATIONAL: w_arb_grant_valid && sched_rd_valid[grant_id]. "
         "Grant is registered and held until ack, so arvalid normally holds "
         "until arready. The live sched_rd_valid AND-term exists to kill "
         "stale grants from the 1-cycle r_arb_request pipeline.",
         "arvalid stable until arready in normal operation; grant one-hot "
         "($onehot0 formal property in-module)",
         f"{RD_ENG}:413. DESIGN NOTE: on the error/reset/timeout paths "
         "sched_rd_valid can drop while an unaccepted AR is pending, "
         "retracting arvalid (AXI stability violation risk) - see K-maps "
         "rd engine sheet."),
        ("Data RD / AR", "m_axi_arid / araddr / arlen", "IW/64/8", "out",
         "axi_read_engine",
         "arid={0,grant_id} (channel in low bits routes R data to the right "
         "SRAM channel). araddr = sched_rd_addr[grant] LIVE from the "
         "scheduler (scheduler advances it only on the done strobe, which "
         "fires on the AR handshake itself). arlen = w_transfer_size = "
         "min(sched_rd_beats-1, cfg_axi_rd_xfer_beats).",
         "sched_rd_addr stable while a grant is pending; arlen+1 <= "
         "sched_rd_beats",
         f"{RD_ENG}:414-419"),
        ("Data RD / AR", "(outstanding gate)", "-", "-", "axi_read_engine",
         "PIPELINE=0: r_outstanding_limit[ch] set on AR handshake, cleared "
         "on rlast handshake; a channel cannot arbitrate while set (one "
         "burst in flight). PIPELINE=1: counter capped at "
         "AR_MAX_OUTSTANDING. Same-cycle AR+rlast handled (independent "
         "ifs, clear wins).",
         "never >1 (or >AR_MAX_OUTSTANDING) outstanding AR per channel",
         f"{RD_ENG}:193-248"),
        ("Data RD / R", "m_axi_rvalid / rready / rdata / rid", "-", "in/out",
         "fabric / axi_read_engine",
         "Pure passthrough to the SRAM controller: sram_valid=rvalid, "
         "sram_data=rdata, sram_id=rid, rready=sram_ready. No internal "
         "buffering - R-channel backpressure IS SRAM backpressure.",
         "beat accepted (rvalid&&rready) == beat written to SRAM ctrl",
         f"{RD_ENG}:469-472"),
        ("Data RD / R", "m_axi_rresp", "2", "in", "fabric",
         "Non-OKAY on an accepted beat sets sched_rd_error[rid] STICKY; "
         "cleared only by external logic (channel reset). Scheduler "
         "escalates to CH_ERROR via w_hard_error.",
         "rresp!=OKAY |=> sched_rd_error[ch] until reset",
         f"{RD_ENG}:520"),
        ("Data WR / AW", "m_axi_awvalid", "1", "out", "axi_write_engine",
         "REGISTERED (r_aw_valid): set when a grant is accepted (grant "
         "valid, no AW pending, live sched_wr_valid), cleared on awready. "
         "Properly holds until the handshake (unlike the read side).",
         "awvalid holds until awready; at most one AW pending in the "
         "register",
         f"{WR_ENG}:562, :576"),
        ("Data WR / AW", "m_axi_awaddr / awlen / awid", "64/8/IW", "out",
         "axi_write_engine",
         "awaddr = sched_wr_addr[r_aw_channel_id] LIVE from the scheduler "
         "(len and channel are latched at grant, the address is NOT). "
         "Contract: the scheduler must hold sched_wr_addr stable while "
         "r_aw_valid is pending - guaranteed because it only advances on "
         "the AW-handshake done strobe.",
         "sched_wr_addr[ch] stable from grant-capture to awready",
         f"{WR_ENG}:579-580"),
        ("Data WR / W", "m_axi_wvalid", "1", "out", "axi_write_engine",
         "r_w_active && sram_valid_reg[ch] && sram_valid_comb[ch]. The "
         "registered valid is the arbitration/timing term; the "
         "combinational valid filters the 1-cycle dry window after the "
         "latency-bridge skid empties, so stale skid data never rides a "
         "wvalid pulse.",
         "wvalid |-> selected channel really has a data beat on "
         "axi_wr_sram_data",
         f"{WR_ENG}:720-722. Dropping the comb term reproduces the dma_2ch "
         "CRC-mismatch (stale-data leak)."),
        ("Data WR / W", "axi_wr_sram_drain (SRAM pop)", "1", "out",
         "axi_write_engine",
         "EXACTLY m_axi_wvalid && m_axi_wready: pop the SRAM unit only on "
         "a real W handshake, so beats consumed == beats transmitted.",
         "SRAM pops == W-channel handshakes, per channel",
         f"{WR_ENG}:704. BUG FIX (deadlock): the previous r_w_active && "
         "wready form popped a beat while wvalid was suppressed; the lost "
         "beat could be the burst's final beat, so WLAST never rode a "
         f"valid beat and the channel hung. See {KI_WLAST}."),
        ("Data WR / W", "m_axi_wlast / wstrb / wuser", "1/DW8/UW", "out",
         "axi_write_engine",
         "wlast = (r_w_beats_remaining==1) - counted down per W handshake "
         "from the shared in-order W-phase FIFO entry. wstrb all-1s "
         "(full-beat writes only). wuser = channel id (instrumentation).",
         "exactly one wlast per awlen+1 beats; W bursts in AW order "
         "(single shared W-phase FIFO)",
         f"{WR_ENG}:725-726"),
        ("Data WR / B", "m_axi_bvalid / bready / bid / bresp", "-", "in/out",
         "fabric / axi_write_engine",
         "bready constant 1 (always ready). bid routes: commit strobe + "
         "beats to the scheduler's commit accumulator, B-phase FIFO pop, "
         "outstanding clear. bresp!=OKAY sets sched_wr_error[bid] sticky.",
         "one B per AW; commit beats == that AW's awlen+1 (B-phase FIFO "
         "entry)",
         f"{WR_ENG}:941, :951"),
        ("Both", "(monitor wrapper)", "-", "-",
         "axi4_master_rd_mon / axi4_master_wr_mon",
         "Both data masters pass through skid+monitor wrappers in "
         "stream_core; the monitor can stall AR/AW via block_ready (see "
         "K-maps core monitors sheet). Sized so this never happens in "
         "normal operation.",
         "RD/WR_MON_MAX_TRANS = NUM_CHANNELS*MAX_OUTSTANDING+4",
         f"{CORE}:73-74. BUG FIX: the previous hardwired 16 compared "
         "against the PER-CHANNEL limit; at >=2 channels the monitor "
         "throttled the datapath."),
    ]
    contract_sheet(
        wb, "Data RD+WR AXI (fub)",
        "STREAM data-path AXI masters (fub side of the skid/monitor "
        "wrappers) signal contracts",
        "Read engine: source memory -> SRAM (AR/R). Write engine: SRAM -> "
        f"destination (AW/W/B). Sources: {RD_ENG}, {WR_ENG}.",
        rows)


def build_apb_csr_contract(wb):
    rows = [
        ("APB kick", "apb_valid / apb_ready / apb_addr", "NCx(1/1/64)",
         "in/out/in", "regfile / descriptor_engine",
         "Per-channel descriptor kick. Accepted ONLY when: addr != 0, no "
         "channel reset, descriptor-address FIFO empty, channel_idle "
         "(scheduler finished all prior work), and no APB op already in "
         "progress (r_apb_ip). apb_addr==0 latches descriptor_error "
         "instead of fetching.",
         "apb handshake |-> channel_idle && addr-FIFO empty && !r_apb_ip; "
         "addr==0 never fetches",
         f"{DESC_ENG}:333-336, :889-891"),
        ("Channel ctrl", "cfg_channel_enable", "NC", "in", "regfile",
         "Gates CH_IDLE -> CH_FETCH_DESC only: a descriptor already in "
         "flight is NOT stopped by dropping enable.",
         "IDLE exit requires descriptor_valid && cfg_channel_enable",
         f"{SCHED}:471"),
        ("Channel ctrl", "cfg_channel_reset", "NC", "in", "regfile",
         "Registered internally; forces the scheduler FSM to CH_IDLE with "
         "top priority, clears beat counters and r_rd_ahead; descriptor "
         "engine aborts in-flight ops and drains to RD_IDLE. The only exit "
         "from sticky CH_ERROR.",
         "reset active |=> FSM==CH_IDLE next cycle; sticky errors clear on "
         "CH_IDLE entry",
         f"{SCHED}:454-456, :797-802"),
        ("Timeout", "cfg_sched_timeout_cycles / _enable / _limit",
         "32/1/8", "in", "regfile",
         "Soft write-progress watchdog: counter runs while sched_wr_valid "
         "&& !sched_wr_ready with no AW-issue and no B-commit; each "
         "elapsed window pulses w_timeout_expired and RE-ARMS. Only "
         "_limit consecutive windows (0 = never) escalate to fatal "
         "CH_ERROR. Any write progress clears the strikes.",
         "bare timeout window never wedges the channel; escalation only "
         "at strikes >= limit",
         f"{SCHED}:1102-1110, :1147-1153. Size cycles to the workload: an "
         "undersized 1000-cycle default falsely timed out 8-channel "
         "contention (top multi_channel hang, fixed by programming the "
         "timeout via APB)."),
        ("Prefetch", "cfg_rd_prefetch_enable", "1", "in", "regfile",
         "Scheduler read-ahead: on a chained LEGACY descriptor the read "
         "side may load descriptor N+1 from the FIFO head (peek, no pop) "
         "while writes still drain N. Capped at ONE ahead. EXT (TASK-101) "
         "descriptors always run lockstep. Clearing it reduces to the "
         "original lockstep FSM verbatim.",
         "r_rd_ahead <= 1 descriptor; disabled => w_rd_peek==w_wr_advance"
         "==0",
         f"{SCHED}:911-927"),
        ("Prefetch", "cfg_prefetch_enable / cfg_fifo_threshold", "1/4",
         "in", "regfile",
         "Descriptor-engine chain throttle: bounds fetched-but-unconsumed "
         "descriptors in the output FIFO. Disabled => on-demand (limit 1); "
         "enabled => up to cfg_fifo_threshold (0 guarded to 1). A "
         "throttled chain fetch is DEFERRED (r_chain_pending), never "
         "dropped.",
         "desc-FIFO occupancy < limit whenever a chain push fires; owed "
         "fetch always issues after drain",
         f"{DESC_ENG}:468-486. BUG FIX: these CSRs were previously "
         "dead (unwired); wiring them + TB timeout sizing fixed the "
         "top-level multi_channel/stress hang."),
        ("Fetch window", "cfg_addr0_base/limit, cfg_addr1_base/limit",
         "4x64", "in", "regfile",
         "Two inclusive address windows; a chained next_descriptor_ptr "
         "must fall inside one of them or the chain stops (no fetch, no "
         "error escalation - engine just goes idle). Also validates the "
         "fetched-from address.",
         "every autonomous fetch address in window0 or window1",
         f"{DESC_ENG}:449-450"),
        ("Engine cfg", "cfg_axi_rd_xfer_beats / cfg_axi_wr_xfer_beats",
         "8/8", "in", "regfile",
         "Stored as the AxLEN VALUE (0 = 1 beat): the engines burst "
         "min(remaining-1, cfg) per transaction. Applies to all channels.",
         "arlen/awlen <= cfg value; final burst may be shorter",
         f"{RD_ENG}:312-313, {WR_ENG}:421-422"),
        ("Monitors", "cfg_desc/rdeng/wreng_mon_enable (+masks)", "-", "in",
         "regfile",
         "Monitor enables are ANDed with the USE_AXI_MONITORS parameter "
         "(tied 0 when the build omits monitors). A DISABLED monitor "
         "never stalls the datapath: the ready gate ORs in "
         "~cfg_monitor_enable.",
         "USE_AXI_MONITORS=0 or cfg off => zero datapath interference",
         f"{CORE}:660/:715; {RD_MON}:477-478"),
    ]
    contract_sheet(
        wb, "APB-CSR config",
        "STREAM configuration / CSR interface contracts (stream_core cfg "
        "ports; register fields live in stream_regs.rdl / stream_mon_regs"
        ".rdl and are accessed BY NAME via the generated regmap)",
        "Every row is checkable against the consuming RTL. Sources: "
        f"{SCHED}, {DESC_ENG}, {CORE}.",
        rows)


def build_monbus_contract(wb):
    rows = [
        ("MonBus", "mon_valid", "1", "out", "scheduler_group_array arbiter",
         "One-cycle PULSE per event packet (state-entry emitters in "
         "scheduler / descriptor_engine, arbitrated up through the group). "
         "Source emitters do NOT wait for mon_ready - a busy consumer "
         "loses the packet at the emitter register.",
         "downstream MUST provide a FIFO (or be always-ready); "
         "mon_valid is never held for backpressure at the source",
         f"{SCHED}:1186-1317 (emitter). Standard integration: gaxi FIFO "
         "or monbus_group SRAM behind the port."),
        ("MonBus", "mon_ready", "1", "in", "consumer",
         "Consumer flow control into the group arbiter. Sustained "
         "backpressure drops packets at the per-source registers, never "
         "corrupts them.",
         "no partial/torn packets under backpressure", ""),
        ("MonBus", "mon_packet", "128", "out", "arbiter",
         "monitor_common_pkg::monitor_packet_t. STREAM events: DESC_START "
         "(CH_FETCH_DESC), DESC_COMPLETE / IRQ (CH_COMPLETE, or in-place "
         "advance under read-ahead), ERROR (CH_ERROR, one-shot), plus "
         "descriptor-engine DESCRIPTOR_LOADED / error packets.",
         "per-descriptor MonBus parity holds with prefetch on or off "
         "(the w_wr_advance path emits the completion the lockstep "
         "CH_COMPLETE path would have)",
         f"{SCHED}:1234-1248"),
        ("MonBus", "mon_timestamp / i_mon_time", "64", "out/in",
         "arbiter / core",
         "Free-running monitor time broadcast; SAMPLED at packet emission "
         "and carried side-band (not inside the 128-bit packet). Held "
         "stable between pulses.",
         "timestamp corresponds to the emit cycle of the packet beside it",
         f"{SCHED}:1209"),
        ("MonBus", "(CH_ERROR one-shot)", "-", "-", "scheduler",
         "CH_ERROR is sticky; the error packet is emitted ONCE per visit "
         "(r_error_pkt_sent latch, cleared on CH_IDLE). Without it the "
         "emitter would flood the trace every cycle.",
         "at most one STREAM_EVENT_ERROR packet per CH_ERROR entry",
         f"{SCHED}:1296-1309"),
        ("MonBus", "(decode discipline)", "-", "-", "DV",
         "Testbenches decode packets via the shared TBClasses.monbus "
         "parser only - never inline bit-shift decodes.",
         "single decoder, split-proof against packet layout changes", ""),
    ]
    contract_sheet(
        wb, "MonBus out",
        "STREAM unified monitor-bus output contracts (stream_core "
        "mon_valid/ready/packet/timestamp)",
        "128-bit packet + 64-bit side-band timestamp. Sources: "
        f"{CORE}:446-450, {SCHED}, {DESC_ENG}.",
        rows)


# ---------------------------------------------------------------------------
# K-map sheets
# ---------------------------------------------------------------------------
def build_rd_engine_kmaps(wb):
    km = new_kmap_sheet(wb, "K-maps rd engine")
    km.sheet_intro(
        f"axi_read_engine - decision K-maps ({RD_ENG})",
        ["Each grid is computed from a python mirror of the exact RTL "
         "expression (file:line cited, RTL quoted). Gray order 00 01 11 10; "
         ">4 vars page into multiple grids.",
         "Comparator sub-terms (space/beat counts) enter the maps as their "
         "1-bit results, factored the way the RTL factors them."])

    km.kmap(
        "w_arb_request[i]", f"{RD_ENG}:327",
        "w_arb_request[i] = sched_rd_valid[i] && w_space_ok[i] && "
        "w_below_outstanding_limit[i]   [below_limit = !r_outstanding_limit]",
        ["sched_rd_valid", "space_ok", "below_limit"],
        lambda v, s, b: v and s and b,
        "EXACTLY ONE 1-cell, at all-ones. A 1 with space_ok=0 over-fills "
        "the SRAM allocation; a 1 with below_limit=0 exceeds the "
        "per-channel outstanding cap. Registered into r_arb_request "
        "(1-cycle pipeline) before the arbiter - stale-grant guards below "
        "exist because of that register.")

    km.table(
        "w_space_ok / w_transfer_size sub-terms", f"{RD_ENG}:312-316",
        ["term", "RTL (mirror of)", "meaning"],
        [("w_transfer_size[i]",
          "(sched_rd_beats <= cfg+1) ? sched_rd_beats-1 : cfg",
          "AxLEN value: full configured burst, or the shorter final burst"),
         ("w_space_ok[i]",
          "axi_rd_alloc_space_free[i] >= w_transfer_size[i]+1",
          "SRAM channel can hold the whole burst BEFORE the AR issues "
          "(pre-allocation; space_free is the registered, 1-cycle-old "
          "view from sram_controller)")],
        note="space_free is registered twice on its way out of the SRAM "
             "controller; the read side tolerates the staleness because "
             "alloc reservations only shrink it (safe direction).")

    km.kmap(
        "m_axi_arvalid", f"{RD_ENG}:413",
        "m_axi_arvalid = w_arb_grant_valid && sched_rd_valid[w_arb_grant_id]",
        ["grant_valid", "sched_rd_valid[gnt]"],
        lambda g, v: g and v,
        "Single 1-cell at (1,1). The sched_rd_valid term masks stale "
        "grants (r_arb_request pipeline). DESIGN NOTE: arvalid is "
        "combinational - if sched_rd_valid drops (error/reset/timeout "
        "escalation) while an AR waits for arready, arvalid retracts, "
        "which violates AXI valid-stability. Benign in the normal flow "
        "(valid only drops via the done-strobe look-ahead, i.e. after the "
        "handshake), but real on the abort paths.")

    km.kmap(
        "w_arb_grant_ack[i]", f"{RD_ENG}:429-431",
        "ack[i] = (grant[i] && m_axi_arvalid && m_axi_arready) | "
        "(grant[i] && !sched_rd_valid[i])   [for the granted channel, "
        "m_axi_arvalid == sched_rd_valid[i]]",
        ["grant_i", "sched_rd_valid_i", "arready"],
        lambda g, v, r: g and ((v and r) or (not v)),
        "1s ONLY in the grant_i=1 half: at (1,1,1) the AR fired, at "
        "(1,0,x) the grant is STALE and auto-releases so the arbiter "
        "advances instead of waiting for an AR that can never fire. The "
        "(1,1,0) cell MUST be 0 - releasing a live grant before arready "
        "would re-arbitrate mid-handshake.")

    km.kmap(
        "r_outstanding_limit[i] next (PIPELINE=0)", f"{RD_ENG}:193-210",
        "set on (arvalid && arready && grant_id==i); clear on (rvalid && "
        "rready && rlast && rid==i); independent ifs, clear written last "
        "so same-cycle AR+RLAST resolves to CLEAR (0)",
        ["ar_fire_i", "rlast_fire_i"],
        lambda a, r: ("clear" if r else ("set" if a else "hold")),
        "clear wins the (1,1) cell - the same-cycle AR/R fix (was "
        "else-if, which wedged the flag at 1). set only at (1,0).",
        values={})

    km.kmap(
        "sched_rd_error[i] latch enable", f"{RD_ENG}:520",
        "latch r_rd_error[rid] on: m_axi_rvalid && m_axi_rready && "
        "(m_axi_rresp != 2'b00)   [resp_ok = (rresp==OKAY)]",
        ["rvalid", "rready", "resp_ok"],
        lambda v, r, ok: v and r and (not ok),
        "Single 1-cell at (1,1,0): only an ACCEPTED bad beat latches the "
        "sticky error. A 1 with rready=0 would latch errors for beats "
        "that were never taken.")

    km.table(
        "R-channel accept / drain wiring (passthrough)",
        f"{RD_ENG}:469-472",
        ["signal", "RTL", "contract"],
        [("axi_rd_sram_valid", "= m_axi_rvalid", "no buffering"),
         ("axi_rd_sram_id", "= m_axi_rid", "channel routing"),
         ("axi_rd_sram_data", "= m_axi_rdata", ""),
         ("m_axi_rready", "= axi_rd_sram_ready",
          "SRAM backpressure is the ONLY R backpressure")],
        note="A beat accepted on R is by construction a beat written to "
             "the SRAM controller - the two counters (dbg_r_beats_rcvd / "
             "dbg_sram_writes) must always match.")


def build_wr_engine_kmaps(wb):
    km = new_kmap_sheet(wb, "K-maps wr engine")
    km.sheet_intro(
        f"axi_write_engine - decision K-maps ({WR_ENG})",
        ["Same conventions as the read-engine sheet. The two bug-fix maps "
         "(axi_wr_sram_drain, m_axi_wvalid) are the ones to re-check after "
         "ANY edit near the W path."])

    km.kmap(
        "w_arb_request[i]", f"{WR_ENG}:452",
        "w_arb_request[i] = sched_wr_valid[i] && w_data_ok[i] && "
        "w_no_outstanding[i]",
        ["sched_wr_valid", "data_ok", "no_outstanding"],
        lambda v, d, n: v and d and n,
        "EXACTLY ONE 1-cell, at all-ones - mirror of the read engine. A 1 "
        "with data_ok=0 issues an AW without backing data (the phantom-"
        "burst class).")

    km.kmap(
        "w_data_ok[i]  (given sched_wr_valid[i])", f"{WR_ENG}:427-434",
        "w_data_ok = w_has_data || w_final_burst ; has_data = "
        "(w_effective_avail >= transfer_size+1) ; final_burst = "
        "(0 < sched_wr_beats <= cfg+1) && (w_effective_avail >= "
        "sched_wr_beats)",
        ["has_data", "final_burst"],
        lambda h, f: h or f,
        "OR of the two comparator results: zero only at (0,0). "
        "final_burst lets the LAST, shorter burst of a descriptor go "
        "with exactly its remaining beats available.")

    km.table(
        "w_effective_avail (stale-view race closure)", f"{WR_ENG}:384-411",
        ["term", "RTL (mirror of)", "why"],
        [("w_drain_t[ch]", "awlen+1 on this cycle's AW handshake",
          "drain_req size firing NOW"),
         ("r_drain_tminus1[ch]", "registered w_drain_t",
          "drain_req that fired LAST cycle"),
         ("w_pending_drain", "r_drain_tminus1 + w_drain_t",
          "in-flight drains not yet visible in the registered avail"),
         ("w_effective_avail", "avail >= pending ? avail - pending : 0",
          "the 'true' count drain_ctrl will reach when our drain lands")],
        note="axi_wr_drain_data_avail is registered twice on its way out "
             "of sram_controller (2-cycle-stale view). Without this "
             "subtraction the engine could fire an AW against beats an "
             "in-flight prior drain_req already claimed -> over-drain -> "
             "permanent occupancy corruption (see K-maps sram ctrl).")

    km.kmap(
        "AW grant-capture enable", f"{WR_ENG}:562",
        "capture if (w_arb_grant_valid && !r_aw_valid && "
        "sched_wr_valid[w_arb_grant_id])",
        ["grant_valid", "r_aw_valid", "sched_wr_valid[gnt]"],
        lambda g, p, v: g and (not p) and v,
        "Single 1-cell at (1,0,1): a new AW is captured only when no AW "
        "is pending and the grant is still live. The (1,0,0) cell MUST "
        "be 0 - that is the stale-grant guard (grant released via "
        "w_stale_grant instead).")

    km.kmap(
        "m_axi_wvalid", f"{WR_ENG}:720-722",
        "m_axi_wvalid = r_w_active && axi_wr_sram_valid[ch] && "
        "axi_wr_sram_valid_comb[ch]",
        ["r_w_active", "valid_reg", "valid_comb"],
        lambda a, vr, vc: a and vr and vc,
        "Single 1-cell at (1,1,1). The (1,1,0) cell is the STALE-SKID "
        "window: registered valid still 1, but the muxed data wire shows "
        "stale skid contents - it MUST be 0 or stale data leaks onto AXI "
        "(the dma_2ch CRC-mismatch shape).")

    km.kmap(
        "axi_wr_sram_drain  (SRAM pop)", f"{WR_ENG}:704",
        "assign axi_wr_sram_drain = m_axi_wvalid && m_axi_wready;",
        ["m_axi_wvalid", "m_axi_wready"],
        lambda v, r: v and r,
        "Single 1-cell at (1,1): pop exactly the beats actually "
        "transmitted. BUG FIX (lost-WLAST deadlock): the pre-fix "
        "'r_w_active && m_axi_wready' form put a 1 in the wvalid=0 "
        "column - the beat was consumed but never transmitted, and when "
        "it was the burst's final beat WLAST never rode a valid beat and "
        "the channel hung (found via the per-channel timing-skew 'mixed' "
        f"profile, kept as regression sentinel). See {KI_WLAST}.")

    km.kmap(
        "w_phase_fifo_pop", f"{WR_ENG}:776-787",
        "pop = (!r_w_active && !fifo_empty) || (m_axi_wvalid && "
        "m_axi_wready && m_axi_wlast && !fifo_empty)   [w_fire = "
        "wvalid && wready]",
        ["r_w_active", "fifo_empty", "w_fire", "wlast"],
        lambda a, e, f, l: ((not a) and (not e)) or (f and l and (not e)),
        "1s only in the fifo_empty=0 columns: start-from-idle (active=0) "
        "and burst-chain-through (w_fire && wlast, no bubble between "
        "back-to-back bursts). Cells with active=0 && w_fire=1 are "
        "unreachable (m_axi_wvalid requires r_w_active).")

    km.kmap(
        "sched_wr_error[i] latch enable", f"{WR_ENG}:951",
        "latch r_wr_error[bid] on: m_axi_bvalid && m_axi_bready && "
        "(m_axi_bresp != 2'b00)   [bready is constant 1]",
        ["bvalid", "resp_ok"],
        lambda v, ok: v and (not ok),
        "Single 1-cell at (1,0). bready==1 always "
        f"({WR_ENG}:941), so bvalid alone qualifies the beat.")


def build_scheduler_kmaps(wb):
    km = new_kmap_sheet(wb, "K-maps scheduler")
    km.sheet_intro(
        f"scheduler - decision K-maps ({SCHED})",
        ["Per-channel FSM: CH_IDLE -> CH_FETCH_DESC -> CH_XFER_DATA -> "
         "CH_COMPLETE -> (CH_NEXT_DESC | CH_IDLE); CH_ERROR sticky. "
         "Flags: rd_done=(r_read_beats_remaining==0), "
         "issue_rem_nz=(r_write_beats_remaining!=0), "
         "commit_zero=(r_write_beats_to_commit==0).",
         "The issue/commit split is the streaming contract: per-descriptor "
         "ADVANCE is issue-based, final channel-done is commit-based."])

    km.kmap(
        "sched_rd_valid  (channel read kick)", f"{SCHED}:962-965",
        "sched_rd_valid = (state==CH_XFER_DATA) && !w_read_complete && "
        "!w_sched_rd_completing_this_cycle && !w_rd_need_base",
        ["state_xfer", "rd_done", "completing_now", "rd_need_base"],
        lambda s, d, c, nb: s and (not d) and (not c) and (not nb),
        "Single 1-cell at (1,0,0,0). completing_now (done-strobe "
        "look-ahead) MUST kill valid the same cycle the final AR's strobe "
        "arrives, or the engine issues one extra transaction against the "
        "not-yet-updated beat counter. rd_need_base stalls reads between "
        "TASK-101 runs.")

    km.kmap(
        "sched_wr_valid  (channel write kick)", f"{SCHED}:988-992",
        "sched_wr_valid = (state==CH_XFER_DATA) && "
        "(r_write_beats_remaining != 0) && !w_write_complete && "
        "!w_sched_wr_completing_this_cycle && !w_wr_need_base",
        ["state_xfer", "issue_rem_nz", "commit_zero",
         "completing_now", "wr_need_base"],
        lambda s, rem, cz, c, nb:
            s and rem and (not cz) and (not c) and (not nb),
        "1s ONLY on the wr_need_base=0 page, in the completing_now=0 "
        "columns, only at (state=1, rem=1, commit_zero=0). The explicit "
        "issue_rem_nz term is load-bearing: w_write_complete tracks "
        "COMMITS now, so without it the engine would keep valid high "
        "through the commit-wait with sched_wr_beats==0 and issue a "
        "garbage AW (transfer-size underflow -> awlen=0xFF). NOTE "
        "commit_zero=1 with rem=1 cannot occur in practice (commits "
        "cannot complete before issue), so that half is unreachable "
        "cover, not a functional path.")

    km.kmap(
        "w_transfer_complete  (per-descriptor advance gate)",
        f"{SCHED}:893",
        "w_transfer_complete = w_read_complete && w_write_issued   "
        "[ISSUE-based on purpose: chained descriptors stream while "
        "their writes commit in the background]",
        ["read_complete", "write_issued"],
        lambda r, w: r and w,
        "Single 1-cell at (1,1). Commit-gating this (as a past edit did) "
        "serializes the write drain at every descriptor boundary - the "
        "regression that motivated the issue/commit split.")

    km.kmap(
        "CH_XFER_DATA exit decision", f"{SCHED}:510-514",
        "if (w_wr_advance) stay-XFER (in-place descriptor advance) ; "
        "else if (w_transfer_complete && !r_rd_ahead) -> CH_COMPLETE ; "
        "else hold",
        ["wr_advance", "transfer_complete", "rd_ahead"],
        lambda a, t, h: ("advance" if a else
                         ("COMPLETE" if (t and not h) else "hold")),
        "advance fills the whole wr_advance=1 half (priority). COMPLETE "
        "only at (0,1,0): while read runs ahead (rd_ahead=1) the "
        "completion exit is SUPPRESSED because w_transfer_complete keys "
        "off the READ counter which already refers to descriptor N+1.",
        values={})

    km.kmap(
        "CH_COMPLETE exit decision", f"{SCHED}:527-531",
        "if (next_ptr != 0 && !last) -> CH_NEXT_DESC (advance NOW, "
        "issue-based) ; else if (w_write_complete) -> CH_IDLE ; else hold",
        ["chained", "write_complete"],
        lambda c, w: ("NEXT_DESC" if c else ("IDLE" if w else "hold")),
        "chained=1 half advances immediately regardless of commits "
        "(streaming). IDLE only at (0,1): the LAST descriptor holds here "
        "until every write has COMMITTED - the channel is never reported "
        "done while data still drains.",
        values={})

    km.kmap(
        "w_rd_peek  (read-ahead loads next descriptor)", f"{SCHED}:916-918",
        "w_rd_peek = w_rd_prefetch_en && w_state_xfer_data && !r_rd_ahead "
        "&& (r_read_beats_remaining==0) && !w_write_issued && "
        "w_desc_chained && descriptor_valid   [head_ok = w_desc_chained "
        "&& descriptor_valid]",
        ["prefetch_en", "state_xfer", "rd_ahead", "rd_done",
         "write_issued", "head_ok"],
        lambda p, s, a, d, w, h:
            p and s and (not a) and d and (not w) and h,
        "1s ONLY on the [write_issued=0, head_ok=1] page, single cell "
        "(prefetch_en=1, state_xfer=1, rd_ahead=0, rd_done=1). rd_ahead=0 "
        "caps read-ahead at ONE descriptor; write_issued=0 means the "
        "write side still owns the FIFO head, so peek must NOT pop it.")

    km.kmap(
        "w_wr_advance  (in-place descriptor advance + FIFO pop)",
        f"{SCHED}:926-927",
        "w_wr_advance = w_rd_prefetch_en && w_state_xfer_data && "
        "w_write_issued && w_desc_chained && descriptor_valid",
        ["prefetch_en", "state_xfer", "write_issued", "chained",
         "desc_valid"],
        lambda p, s, w, c, v: p and s and w and c and v,
        "1s ONLY on the [desc_valid=1] page, single cell at all-ones. "
        "This is the ONLY XFER-state descriptor_ready source: it pops "
        "the FIFO head that peek borrowed. prefetch_en=0 forces the "
        "whole map to 0 (lockstep A/B on one bitstream).")

    km.kmap(
        "descriptor_ready  (FIFO pop)", f"{SCHED}:1064-1066",
        "descriptor_ready = (state==CH_IDLE) || (state==CH_NEXT_DESC) || "
        "w_wr_advance",
        ["state_idle", "state_next_desc", "wr_advance"],
        lambda i, n, a: i or n or a,
        "OR of three sources; zero only at all-zeros. idle/next_desc are "
        "mutually exclusive (one-hot FSM) and wr_advance can only be 1 in "
        "CH_XFER_DATA, so at most one source is live per cycle - the "
        "(1,1,x) cells are unreachable cover.")

    km.kmap(
        "w_hard_error  (fatal -> sticky CH_ERROR)", f"{SCHED}:1157-1158",
        "w_hard_error = descriptor_error || sched_rd_error || "
        "sched_wr_error || r_read_error_sticky || r_write_error_sticky   "
        "[rd_any = live|sticky, wr_any = live|sticky]",
        ["descriptor_error", "rd_any", "wr_any"],
        lambda d, r, w: d or r or w,
        "Zero ONLY at all-zeros (pure OR). A bare timeout window is "
        "deliberately NOT in this cone - it escalates separately via "
        "w_timeout_escalate (strikes >= cfg_sched_timeout_limit).")

    km.table(
        "r_timeout_counter next-value (priority order)",
        f"{SCHED}:1102-1110",
        ["condition (first match wins)", "next", "meaning"],
        [("sched_wr_done_strobe || sched_wr_commit_strobe", "0",
          "ANY write progress (AW issue OR B commit) re-arms - commits "
          "count as progress because completion is commit-gated"),
         ("w_timeout_expired", "0",
          "window elapsed -> re-arm and keep waiting (SOFT timeout, one "
          "pulse per window)"),
         ("sched_wr_valid && !sched_wr_ready", "counter + 1",
          "genuinely waiting on the write engine"),
         ("otherwise", "0", "not waiting")],
        note="w_timeout_expired = cfg_sched_timeout_enable && (counter >= "
             f"cfg_sched_timeout_cycles) ({SCHED}:1147-1148). Escalation: "
             "w_timeout_escalate = (limit != 0) && (strikes >= limit) "
             f"({SCHED}:1152-1153); strikes clear on any progress or "
             "CH_IDLE.")

    km.table(
        "FSM transition priority", f"{SCHED}:451-465",
        ["#", "condition", "action"],
        [(1, "r_channel_reset_active", "-> CH_IDLE (overrides all)"),
         (2, "w_hard_error || w_timeout_escalate", "-> CH_ERROR (sticky)"),
         (3, "state case (normal transitions)", "see exit-decision maps")],
        note="CH_ERROR self-loops until channel reset; scheduler_idle "
             "reports CH_IDLE only (a wedged channel must NOT read idle).")


def build_desc_engine_kmaps(wb):
    km = new_kmap_sheet(wb, "K-maps desc engine")
    km.sheet_intro(
        f"descriptor_engine - decision K-maps ({DESC_ENG})",
        ["Fetch FSM: RD_IDLE -> RD_ISSUE_ADDR -> RD_WAIT_DATA "
         "(-> RD_ISSUE_ADDR2 -> RD_WAIT_DATA2 for EXT) -> RD_COMPLETE; "
         "RD_ERROR on bad response. Chaining pushes next_descriptor_ptr "
         "back into the address FIFO, throttled by the prefetch limit."])

    km.kmap(
        "w_chain_condition  (level 1: basic eligibility)",
        f"{DESC_ENG}:454",
        "w_chain_condition = (w_next_addr != '0) && !w_desc_last && "
        "w_desc_valid",
        ["next_addr_nz", "last", "valid"],
        lambda n, l, v: n and (not l) and v,
        "Single 1-cell at (1,0,1). last=1 overrides a non-zero pointer "
        "(explicit termination); valid=0 never chains.")

    km.kmap(
        "w_chain_eligible  (level 2+3, throttle-independent)",
        f"{DESC_ENG}:457-459",
        "w_chain_eligible = w_chain_condition && w_next_addr_valid && "
        "!r_descriptor_error   [next_addr_valid = in cfg_addr0 or "
        "cfg_addr1 window]",
        ["chain_condition", "addr_in_window", "descriptor_error"],
        lambda c, a, e: c and a and (not e),
        "Single 1-cell at (1,1,0). The window check is the runaway-chain "
        "guard: an out-of-window pointer silently ENDS the chain (engine "
        "idles; no error escalation).")

    km.kmap(
        "w_should_chain  (immediate chain push)", f"{DESC_ENG}:479-480",
        "w_should_chain = w_chain_eligible && w_desc_committed && "
        "w_prefetch_allows && w_desc_addr_fifo_wr_ready   [committed = "
        "(state==RD_COMPLETE) && desc-FIFO wr_ready]",
        ["eligible", "committed", "prefetch_allows", "addr_fifo_ready"],
        lambda e, c, p, f: e and c and p and f,
        "Single 1-cell at all-ones. eligible && committed && !push "
        "(throttled or FIFO full) does NOT lose the fetch - it arms "
        "r_chain_pending (deferred path below).")

    km.kmap(
        "w_pending_push_fire  (deferred chain push)", f"{DESC_ENG}:485-486",
        "w_pending_push_fire = r_chain_pending && w_prefetch_allows && "
        "w_desc_addr_fifo_wr_ready && !w_desc_committed",
        ["chain_pending", "prefetch_allows", "addr_fifo_ready",
         "committed"],
        lambda p, a, f, c: p and a and f and (not c),
        "Single 1-cell at (1,1,1,0). The !committed term keeps the "
        "commit cycle reserved for the immediate path - both pushing in "
        "one cycle would double-write the address FIFO mux.")

    km.table(
        "w_prefetch_limit  (chain throttle)", f"{DESC_ENG}:468-476",
        ["cfg_prefetch_enable", "cfg_fifo_threshold", "limit",
         "behavior"],
        [("0", "x", "1", "on-demand: fetch next only after the scheduler "
          "drains the current descriptor"),
         ("1", "0", "1", "guard: threshold 0 would stall chaining "
          "forever"),
         ("1", "N>0", "N", "buffer up to N fetched-but-unconsumed "
          "descriptors ahead of the scheduler")],
        note="w_prefetch_allows = (w_desc_fifo_count < w_prefetch_limit) "
             f"({DESC_ENG}:476). BUG FIX: cfg_prefetch_enable / "
             "cfg_fifo_threshold were previously DEAD (unwired) - wiring "
             "them into this throttle was half of the top-level "
             "multi_channel/stress hang fix.")

    km.kmap(
        "APB accept  (w_apb_skid_valid_in / apb_ready)",
        f"{DESC_ENG}:333-336",
        "accept = apb_valid && !r_channel_reset_active && "
        "w_desc_addr_fifo_empty && channel_idle && !r_apb_ip   "
        "[apb_ready has the same gate off the skid ready]",
        ["apb_valid", "reset_active", "addr_fifo_empty", "channel_idle",
         "apb_ip"],
        lambda v, r, e, i, p: v and (not r) and e and i and (not p),
        "1s ONLY on the apb_ip=0 page, in the channel_idle=1 columns, "
        "single cell (valid=1, reset=0, fifo_empty=1). fifo_empty "
        "prevents APB from "
        "clobbering a queued chain address; apb_ip prevents accepting a "
        "second kick before the first chain completes (cleared on the "
        "channel_idle falling edge).")

    km.kmap(
        "ar_valid  (descriptor fetch AR)", f"{DESC_ENG}:920-921",
        "ar_valid = ((state==RD_ISSUE_ADDR) || (state==RD_ISSUE_ADDR2)) "
        "&& !r_axi_read_active",
        ["st_issue_addr", "st_issue_addr2", "read_active"],
        lambda s1, s2, a: (s1 or s2) and (not a),
        "1s only in the read_active=0 half. The (1,1,x) cells are "
        "unreachable (one-hot FSM). read_active latches on ar_ready and "
        "clears on the R response, so a state re-entry cannot double-"
        "issue the same AR.")

    km.kmap(
        "descriptor_valid  (to scheduler)", f"{DESC_ENG}:1037",
        "descriptor_valid = w_desc_fifo_rd_valid && !r_descriptor_error",
        ["fifo_rd_valid", "descriptor_error"],
        lambda v, e: v and (not e),
        "Single 1-cell at (1,0): a latched fetch error blocks ALL "
        "buffered descriptors from reaching the scheduler until the "
        "error clears (RD_IDLE re-entry / channel reset).")


def build_sram_kmaps(wb):
    km = new_kmap_sheet(wb, "K-maps sram ctrl")
    km.sheet_intro(
        "sram_controller / stream_alloc_ctrl / stream_drain_ctrl - "
        f"accounting K-maps ({SRAM_UNIT}, {ALLOC}, {DRAIN})",
        ["Two virtual FIFOs of POINTERS bracket the real per-channel data "
         "FIFO: alloc_ctrl tracks RESERVED-vs-landed space (read engine "
         "side), drain_ctrl tracks LANDED-vs-reserved data (write engine "
         "side). All engine-visible counts derive from these pointers."])

    km.kmap(
        "alloc reserve advance  (wr_ptr += size)",
        f"{ALLOC}:76-92 via {SRAM_UNIT}:134",
        "advance = wr_valid && wr_ready && !r_wr_full   [wr_valid = "
        "axi_rd_alloc_req (fires 1 cycle after each AR handshake); "
        "wr_ready = !full]",
        ["alloc_req", "wr_full"],
        lambda q, f: q and (not f),
        "Single 1-cell at (1,0). The reservation is made AFTER the AR "
        "already issued - safety comes from the read engine's w_space_ok "
        "pre-check, not from this full flag.")

    km.kmap(
        "alloc release  (rd_ptr += 1, space freed)", f"{SRAM_UNIT}:141",
        ".rd_valid(axi_wr_sram_valid && axi_wr_sram_ready)   [release "
        "fires on the BRIDGE-OUTPUT handshake, i.e. when a beat leaves "
        "the channel unit toward the write engine]",
        ["bridge_out_valid", "bridge_out_ready"],
        lambda v, r: v and r,
        "Single 1-cell at (1,1). Space is freed at unit EXIT (not FIFO "
        "entry): a beat parked in the latency-bridge skid still occupies "
        "its reservation, so space_free can never over-report.")

    km.table(
        "space_free accounting chain (read side)",
        f"{ALLOC}:141, {SRAM_UNIT}:310-316, {SRAM_CTL}:257-267",
        ["stage", "expression / register", "staleness"],
        [("alloc_ctrl", "space_free = DEPTH - (wr_ptr - rd_ptr)",
          "combinational"),
         ("sram_controller_unit", "registered once",
          "1 cycle"),
         ("sram_controller wrapper", "registered again "
          "(axi_rd_alloc_space_free port)", "2 cycles total"),
         ("axi_read_engine", "w_space_ok compares the stale view",
          "safe: reservations only SHRINK the stale view")],
        note="The 2-flop stage is the 100 MHz timing closure for the "
             "8-channel arbitration cone; false-pathing it was rejected "
             "because a stale grant handshake could latch the wrong "
             "channel.")

    km.kmap(
        "drain occupancy advance  (wr_ptr += 1)", f"{SRAM_UNIT}:184",
        ".wr_valid(axi_rd_sram_valid && axi_rd_sram_ready)   [data beat "
        "lands in the channel FIFO -> data_available += 1]",
        ["rd_sram_valid", "rd_sram_ready"],
        lambda v, r: v and r,
        "Single 1-cell at (1,1): occupancy counts only ACCEPTED beats.")

    km.kmap(
        "drain reserve advance  (rd_ptr += rd_size)", f"{DRAIN}:97-105",
        "advance = rd_valid && rd_ready && !r_rd_empty   [rd_valid = "
        "axi_wr_drain_req (fires on AW handshake); rd_ready = !empty; "
        "advances by the FULL rd_size]",
        ["drain_req", "rd_empty"],
        lambda q, e: q and (not e),
        "Single 1-cell at (1,0). DANGER cell: the advance is gated only "
        "on not-EMPTY, not on size<=available - a reservation larger "
        "than the real occupancy makes rd_ptr overshoot wr_ptr and the "
        "wrap-corrected count reads ~DEPTH FOREVER (permanent corruption, "
        "all-channel freeze). A sim-only $error at "
        f"{DRAIN}:172-181 fires at the bad reservation; the write "
        "engine's w_effective_avail pipeline is the synthesis-side "
        "guarantee the reservation is never oversized.")

    km.table(
        "axi_wr_drain_data_avail contract (write side)",
        f"{SRAM_UNIT}:282-307",
        ["rule", "detail"],
        [("avail = drain_ctrl occupancy ONLY",
          "assign axi_wr_drain_data_avail = drain_data_available; "
          "(w_count of the drain virtual FIFO)"),
         ("bridge_occupancy is EXCLUDED",
          "a beat in the latency-bridge skid is still counted by "
          "drain_ctrl (its rd_ptr only moves on drain_req), so adding "
          "bridge_occupancy counted it TWICE"),
         ("registered once more at the wrapper",
          f"{SRAM_CTL}:257-267 -> the write engine sees a 2-cycle-stale "
          "view and compensates with w_effective_avail")],
        note="BUG FIX (all-channel deadlock): the double-count let the "
             "write engine size a burst against beats that did not "
             "exist -> over-drain -> permanent occupancy corruption -> "
             "stall at the head of the shared in-order W-phase FIFO "
             f"froze all 8 channels (slow_producer profile). See "
             f"{KI_DBLCNT}. Excluding skid beats is conservative-safe: "
             "it can only delay an AW, never over-issue one.")


def build_core_monitor_kmaps(wb):
    km = new_kmap_sheet(wb, "K-maps core monitors")
    km.sheet_intro(
        f"stream_core monitor gating + block_ready ({CORE}, {RD_MON}, "
        f"{WR_MON}, {MON_BASE})",
        ["The data-path AXI masters run through axi4_master_rd_mon / "
         "axi4_master_wr_mon wrappers. The monitor's block_ready is ANDed "
         "into the fub-side AR/AW ready, so a saturated monitor table "
         "THROTTLES the engines instead of losing transactions - and the "
         "saturation-recovery contract guarantees it un-throttles."])

    km.kmap(
        "fub_axi_arready gate  (read datapath)", f"{RD_MON}:477-478",
        "fub_axi_arready = w_core_fub_axi_arready & (w_block_ready | "
        "~cfg_monitor_enable)   [identical shape for AW: "
        f"{WR_MON}:482-483]",
        ["core_ready", "block_ready", "mon_enable"],
        lambda c, b, e: c and (b or (not e)),
        "Zero only where core_ready=0 or (block_ready=0 AND "
        "mon_enable=1). The mon_enable=0 column must equal core_ready "
        "verbatim - a DISABLED monitor never stalls the datapath. A 0 at "
        "(1,1,1) would be a monitor stalling while claiming capacity.")

    km.kmap(
        "w_block_ready  (monitor capacity)", f"{MON_BASE}:483-485",
        "block_ready = (MAX_TRANSACTIONS > BLOCK_MARGIN) ? (active_count "
        "< MAX_TRANSACTIONS - BLOCK_MARGIN) : 1'b1   [vars: max_gt_margin "
        "(param, compile-time), count_below (comparator)]",
        ["max_gt_margin", "count_below"],
        lambda m, c: (c if m else True),
        "block_ready follows count_below whenever the table is big "
        "enough to carry a margin; degenerate tables never block. "
        "POLARITY is the loaded gun here: 1 = proceed. The inverted "
        "pre-fix polarity deadlocked every upstream handshake at reset.")

    km.table(
        "saturation-recovery contract (stream default sizing)",
        f"{MON_BASE}:456-485, {MON_PKG}:114-116, {CORE}:73-74",
        ["quantity", "expression", "8ch default value"],
        [("MAX_TRANSACTIONS", "NUM_CHANNELS * AR_MAX_OUTSTANDING + 4",
          "68"),
         ("cmd_entry_reserve(MAX)", "(MAX >= 16) ? 2 : 0", "2"),
         ("command-entry cap (trans_mgr)", "MAX - reserve", "66"),
         ("BLOCK_MARGIN", "reserve - 1 (or flat 3 if reserve==0)", "1"),
         ("block threshold", "block when active_count >= MAX - margin",
          ">= 67"),
         ("reopen threshold", "re-assert when active_count < MAX - margin",
          "< 67")],
        note="The reopen threshold sits STRICTLY ABOVE the command cap "
             "(67 > 66): even a table whose command entries are all "
             "permanently in flight recovers block_ready as soon as "
             "orphan entries drain. The old flat MAX-3 margin parked the "
             "table exactly AT the threshold and never re-asserted "
             "(stream_core multi-channel wedge). Undersizing "
             "MAX_TRANSACTIONS on purpose (mon_small config) therefore "
             "throttles-and-recovers, never deadlocks.")

    km.table(
        "monitor enable tie-off (build-level gate)", f"{CORE}:641-732",
        ["USE_AXI_MONITORS", "int_cfg_*_mon_enable", "effect"],
        [("1", "= cfg_*_mon_enable (CSR)", "runtime on/off; wrapper "
          "instantiates the monitor (USE_MONITOR=1)"),
         ("0", "= 1'b0 (tied)", "monitor logic absent, w_block_ready "
          "tied 1: bare skid, zero datapath interference")],
        note=f"e.g. {CORE}:660 (enabled branch) vs :715 (tied-off "
             "branch). The always-on axi_bus_meter perf buckets survive "
             "USE_AXI_MONITORS=0 by design - only the heavy monitors and "
             "histograms are gated.")


def main():
    verify_citations()
    wb = openpyxl.Workbook()
    del wb[wb.sheetnames[0]]           # drop the default sheet

    build_desc_axi_contract(wb)
    build_data_axi_contract(wb)
    build_apb_csr_contract(wb)
    build_monbus_contract(wb)

    build_rd_engine_kmaps(wb)
    build_wr_engine_kmaps(wb)
    build_scheduler_kmaps(wb)
    build_desc_engine_kmaps(wb)
    build_sram_kmaps(wb)
    build_core_monitor_kmaps(wb)

    wb.save(XLSX)
    print(f"wrote {XLSX}")
    for name in wb.sheetnames:
        print(" ", name)


if __name__ == "__main__":
    main()
