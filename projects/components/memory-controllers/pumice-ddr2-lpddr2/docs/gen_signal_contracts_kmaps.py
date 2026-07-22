#!/usr/bin/env python3
# SPDX-License-Identifier: MIT
# SPDX-FileCopyrightText: 2026 sean galloway
"""Update pumice_signal_contracts.xlsx for the current RTL + K-map pages.

Adds/updates:
  * stale "SUSPECT" notes from the read-debug era -> RESOLVED pointers
  * scheduler-sheet rows the RTL grew since (refresh iface, t_rfc_i, checker)
  * K-map sheets for every important decision signal in the arbiter, the bank
    timer and the refresh path. Each map is computed from a python mirror of
    the EXACT RTL expression (file:line referenced), so grid and RTL agree by
    construction; the "inspection check" note says what a healthy map looks
    like so a bad edit shows up on sight.

Gray order everywhere: 00 01 11 10. >4 variables = one 4x4 grid per value of
the extra (page) variables. 1-cells green, 0-cells grey.

Rerun after arbiter/bank-timer/refresh RTL changes:
    python3 docs/gen_signal_contracts_kmaps.py
"""
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "pumice_signal_contracts.xlsx")

GREEN = PatternFill("solid", fgColor="C6EFCE")
GREY = PatternFill("solid", fgColor="F2F2F2")
TITLE = Font(bold=True, size=12)
HDR = Font(bold=True)
MONO = Font(name="Consolas", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Border(*[Side(style="thin")] * 4)

GRAY2 = [(0, 0), (0, 1), (1, 1), (1, 0)]  # gray-code order of 2 bits


def _glabel(bits):
    return "".join(str(b) for b in bits)


class KmapWriter:
    def __init__(self, ws):
        self.ws = ws
        self.row = 1

    def sheet_intro(self, title, lines):
        ws = self.ws
        ws.cell(self.row, 1, title).font = TITLE
        self.row += 1
        for ln in lines:
            c = ws.cell(self.row, 1, ln)
            c.alignment = WRAP
            self.row += 1
        self.row += 1

    def kmap(self, name, source, expr, varnames, fn, check, values=None):
        """One K-map block. varnames: MSB-first list (2..6). fn(*bits)->0/1
        (or a short string when `values` mapping is wanted). Pages over
        varnames[4:]."""
        ws = self.ws
        ws.cell(self.row, 1, name).font = TITLE
        ws.cell(self.row, 4, source).font = MONO
        self.row += 1
        c = ws.cell(self.row, 1, expr)
        c.font = MONO
        c.alignment = WRAP
        self.row += 1
        n = len(varnames)
        rowv = varnames[0:min(2, n)]                  # grid row variables
        colv = varnames[min(2, n):min(4, n)]          # grid col variables
        pagev = varnames[4:]                          # page variables
        ws.cell(self.row, 1,
                f"rows = {'/'.join(rowv)}   cols = {'/'.join(colv)}"
                + (f"   pages = {'/'.join(pagev)}" if pagev else ""))
        self.row += 1
        ws.cell(self.row, 1, f"CHECK BY INSPECTION: {check}").alignment = WRAP
        ws.cell(self.row, 1).font = Font(italic=True)
        self.row += 2

        rows = GRAY2 if len(rowv) == 2 else ([(0,), (1,)] if len(rowv) == 1
                                             else [()])
        cols = GRAY2 if len(colv) == 2 else ([(0,), (1,)] if len(colv) == 1
                                             else [()])
        pages = [()]
        for _ in pagev:
            pages = [p + (b,) for p in pages for b in (0, 1)]

        for page in pages:
            base = self.row
            if pagev:
                lbl = ", ".join(f"{v}={b}" for v, b in zip(pagev, page))
                ws.cell(base, 1, f"[{lbl}]").font = HDR
                base += 1
            # column headers
            for j, cb in enumerate(cols):
                cc = ws.cell(base, 2 + j, _glabel(cb) or "-")
                cc.font = HDR
                cc.alignment = CENTER
            for i, rb in enumerate(rows):
                rc = ws.cell(base + 1 + i, 1, _glabel(rb) or "-")
                rc.font = HDR
                rc.alignment = CENTER
                for j, cb in enumerate(cols):
                    bits = tuple(rb) + tuple(cb) + tuple(page)
                    v = fn(*bits)
                    cell = ws.cell(base + 1 + i, 2 + j)
                    if values:                        # multi-valued map
                        cell.value = values.get(v, str(v))
                        benign = str(v) in ("0", "-", "wait", "hold", "IDLE")
                        cell.fill = GREY if benign else GREEN
                    else:
                        cell.value = int(bool(v))
                        cell.fill = GREEN if v else GREY
                    cell.alignment = CENTER
                    cell.border = THIN
            self.row = base + 1 + len(rows) + 1
        self.row += 1

    def table(self, name, source, headers, rows, note=""):
        ws = self.ws
        ws.cell(self.row, 1, name).font = TITLE
        ws.cell(self.row, 4, source).font = MONO
        self.row += 1
        if note:
            c = ws.cell(self.row, 1, note)
            c.alignment = WRAP
            c.font = Font(italic=True)
            self.row += 1
        for j, h in enumerate(headers):
            c = ws.cell(self.row, 1 + j, h)
            c.font = HDR
            c.border = THIN
        self.row += 1
        for r in rows:
            for j, v in enumerate(r):
                c = ws.cell(self.row, 1 + j, v)
                c.border = THIN
                c.alignment = WRAP
            self.row += 1
        self.row += 2


def build_arbiter_sheet(wb):
    if "K-maps arbiter" in wb.sheetnames:
        del wb["K-maps arbiter"]
    ws = wb.create_sheet("K-maps arbiter")
    for col, w in zip("ABCDEFGH", (26, 9, 9, 9, 9, 9, 9, 9)):
        ws.column_dimensions[col].width = w
    km = KmapWriter(ws)
    km.sheet_intro(
        "pumice_cmd_arbiter — decision K-maps (rtl/fub/pumice_cmd_arbiter.sv)",
        ["Each grid is computed from a python mirror of the exact RTL "
         "expression. Variables are the 1-bit conditions feeding the decision "
         "(counter==0 flags etc.). Gray order 00 01 11 10; >4 vars page into "
         "multiple grids.",
         "Healthy shapes are stated per map — a regression that opens an "
         "illegal cell is visible by inspection."])

    # 1. w_ref_safe (guards collapsed to one flag: guard0|guard1 nonzero)
    km.kmap(
        "w_ref_safe", "pumice_cmd_arbiter.sv (refresh pick gate)",
        "w_ref_safe = !w_any_active && !w_inflight_preact && (r_guard0=='0) && "
        "(r_guard1=='0) && !w_rfc_busy   [guards collapsed: guards_nz = "
        "|r_guard0 | |r_guard1]",
        ["any_active", "inflight_preact", "guards_nz", "rfc_busy"],
        lambda a, i, g, r: (not a) and (not i) and (not g) and (not r),
        "EXACTLY ONE 1-cell, at all-zeros. Any additional 1 is a hole that "
        "lets REFab collide with an open/opening row or violate tRFC "
        "(the silicon row-corruption bug class).")

    # 2. refresh branch action (multi-valued)
    km.kmap(
        "refresh-branch action (given refresh_req_i||refresh_drain_i)",
        "pumice_cmd_arbiter.sv (priority 2)",
        "if (any_active) { if (rfsh_pre_found) PRE else wait } "
        "else if (ref_safe) REF else wait",
        ["any_active", "rfsh_pre_found", "ref_safe"],
        lambda a, f, s: ("PRE" if (a and f) else
                         ("REF" if ((not a) and s) else "wait")),
        "REF appears ONLY where any_active=0 AND ref_safe=1. PRE only where "
        "any_active=1 AND pre_found=1. Everything else waits (the branch "
        "never falls through to column/ACT picks).",
        values={})

    # 3. column masks (6 vars -> 4 pages)
    km.kmap(
        "rd_col_m[e]  (given rd_sch_valid_i[e] && rd_issue_ready)",
        "pumice_cmd_arbiter.sv (classify + direction guard)",
        "rd_col_m = rhit && bank_rdwr_ready && tccd_ok && twtr_ok && "
        "rd_issue_ready && !w_inflight_col && !w_rd_turn_block   "
        "[rd_issue_ready factored out as the enabling condition]",
        ["rhit", "rdwr_ready", "tccd_ok", "twtr_ok",
         "inflight_col", "rd_turn_block"],
        lambda h, r, c, w, f, t: h and r and c and w and (not f) and (not t),
        "1s ONLY on the page [inflight_col=0, rd_turn_block=0], single "
        "all-ones cell. ANY 1 on an rd_turn_block=1 page = a RD issued into "
        "a write burst's DQ occupancy on the stale flopped twtr_ok — the "
        "471/471 concurrent-soak corruption (issue #42).")
    km.kmap(
        "wr_col_m[e]  (given wr_sch_valid_i[e] && wr_commit_ready)",
        "pumice_cmd_arbiter.sv (classify + direction guard)",
        "wr_col_m = whit && bank_rdwr_ready && tccd_ok && trtw_ok && "
        "wr_commit_ready && !w_inflight_col && !w_wr_turn_block   "
        "[wr_commit_ready factored out as the enabling condition]",
        ["whit", "rdwr_ready", "tccd_ok", "trtw_ok",
         "inflight_col", "wr_turn_block"],
        lambda h, r, c, t, f, b: h and r and c and t and (not f) and (not b),
        "Mirror of rd_col_m: 1s only on [inflight_col=0, wr_turn_block=0].")
    km.kmap(
        "w_rd_turn_block / w_wr_turn_block",
        "pumice_cmd_arbiter.sv (direction-turnaround guard)",
        "w_rd_turn_block = r_wrfire0 || r_wrfire1 ; "
        "w_wr_turn_block = r_rdfire0 || r_rdfire1   (fire-shift of the "
        "OPPOSITE direction; covers the 2 cycles until the flopped "
        "twtr/trtw ok reflects the fired column)",
        ["oppfire0", "oppfire1"],
        lambda f0, f1: f0 or f1,
        "Zero ONLY at (0,0). The guard is direction-CROSSED: a fired WR "
        "blocks RD picks and vice versa; same-direction pacing stays with "
        "tCCD. If either 1-cell reads 0, the turnaround hole is back.")

    # 4. activate masks (6 vars)
    km.kmap(
        "rd_act_m[e] / wr_act_m[e]  (given sch_valid[e])",
        "pumice_cmd_arbiter.sv (classify + tRFC gate)",
        "act_m = !bank_row_active && !w_guarded[bank] && bank_act_ready && "
        "tfaw_ok && trrd_ok && !w_rfc_busy",
        ["row_active", "guarded", "act_ready", "tfaw_ok",
         "trrd_ok", "rfc_busy"],
        lambda ra, g, ar, tf, tr, rb:
            (not ra) and (not g) and ar and tf and tr and (not rb),
        "1s ONLY on the page [trrd_ok=1, rfc_busy=0], single cell "
        "(row_active=0, guarded=0, act_ready=1, tfaw_ok=1). ANY 1 on an "
        "rfc_busy=1 page = ACT during refresh recovery — the silicon "
        "row-corruption bug the tRFC counter closes.")

    # 5. precharge masks (4 vars, clean single grid)
    km.kmap(
        "rd_pre_m[e] / wr_pre_m[e]  (given sch_valid[e])",
        "pumice_cmd_arbiter.sv (classify)",
        "pre_m = bank_row_active && !w_guarded[bank] && !hit && "
        "bank_pre_ready",
        ["row_active", "guarded", "hit", "pre_ready"],
        lambda ra, g, h, p: ra and (not g) and (not h) and p,
        "Single 1-cell at (1,0,0,1): only an open bank on the WRONG row, "
        "un-guarded and tRAS/tRTP/tWR-clear, may precharge. A 1 with "
        "guarded=1 = the registered-readiness staleness hazard.")

    # 6. per-bank guard
    km.kmap(
        "w_guarded[b]",
        "pumice_cmd_arbiter.sv (guard fold)",
        "w_guarded[b] = r_guard0[b] | r_guard1[b] | ((w_inflight_preact || "
        "w_inflight_col) && (r_bank == b))   [bank_match = r_bank==b]",
        ["guard0", "guard1", "inflight_rowop_or_col", "bank_match"],
        lambda g0, g1, i, m: g0 or g1 or (i and m),
        "Zero ONLY when both guard stages are clear AND no in-flight "
        "row-affecting/column op targets this bank. Columns are included "
        "(tRTP/tWR registration lag) — if the (0,0,1,1) cell ever reads 0, "
        "the column-guard extension was lost.")

    # 7. output stage
    km.kmap(
        "w_out_ready / w_fire_out",
        "pumice_cmd_arbiter.sv (output register)",
        "w_out_ready = !r_pick_valid || cmd_ready_i ; "
        "w_fire_out = r_pick_valid && cmd_ready_i",
        ["pick_valid", "cmd_ready"],
        lambda p, c: ("rdy+fire" if (p and c) else
                      ("rdy" if not p else "hold")),
        "hold ONLY at (1,0) — a full cmd FIFO holds the decision; "
        "fire ONLY at (1,1). Guards/commits/grants strobe on fire alone.",
        values={})

    # 8. priority order table
    km.table(
        "pick priority (strict, first match wins)",
        "pumice_cmd_arbiter.sv (always_comb pick)",
        ["#", "condition", "action", "note"],
        [(1, "!init_done_i && init_cmd_valid_i", "forward init op",
          "MRS/PRE/REF from init_sequencer verbatim"),
         (2, "refresh_req_i || refresh_drain_i", "refresh branch",
          "see refresh-branch K-map; never falls through while pending"),
         (3, "rd_col_f", "RD/RDA row-hit", "read priority over write"),
         (4, "wr_col_f", "WR/WRA row-hit", ""),
         (5, "rd_act_f", "ACT for oldest pending read", "bank-parallel"),
         (6, "wr_act_f", "ACT for oldest pending write", ""),
         (7, "rd_pre_f", "PRE wrong-row bank (read)", ""),
         (8, "wr_pre_f", "PRE wrong-row bank (write)", "")],
        note="Strict if/else cascade — exactly one action per cycle; refresh "
             "starves columns by design while req is high (liveness relies "
             "on tREFI >> refresh service time).")

    # 9. tRFC counter rule
    km.table(
        "r_rfc_cnt next-value", "pumice_cmd_arbiter.sv (tRFC counter)",
        ["w_fire_out && r_grant (REF fired)", "r_rfc_cnt != 0", "next"],
        [("1", "x", "t_rfc_i (load)"),
         ("0", "1", "r_rfc_cnt - 1"),
         ("0", "0", "0 (idle)")],
        note="w_rfc_busy = (r_rfc_cnt != 0). Load wins over decrement; "
             "blocks ACT picks and further REFs (drain REFs are tRFC-spaced).")
    return ws


def build_bank_timer_sheet(wb):
    if "K-maps bank timer" in wb.sheetnames:
        del wb["K-maps bank timer"]
    ws = wb.create_sheet("K-maps bank timer")
    for col, w in zip("ABCDEFGH", (26, 9, 9, 9, 9, 9, 9, 9)):
        ws.column_dimensions[col].width = w
    km = KmapWriter(ws)
    km.sheet_intro(
        "bank_timer — safe-signal K-maps (rtl/fub/bank_timer.sv)",
        ["Flags: rv=r_row_valid, ap=r_ap_pending, X0=(timer X == 0). "
         "safe_* are pure combinational ANDs over these — one register stage "
         "(the counters) behind, which is why the arbiter carries the 2-cycle "
         "guard."])

    km.kmap(
        "safe_act_o", "bank_timer.sv:133",
        "safe_act = !r_row_valid && (r_rp == 0) && (r_rc == 0)",
        ["row_valid", "rp==0", "rc==0"],
        lambda rv, rp0, rc0: (not rv) and rp0 and rc0,
        "Single 1-cell at (0,1,1): closed row, tRP and tRC elapsed. Any 1 "
        "with row_valid=1 would re-ACT an open bank.")
    km.kmap(
        "safe_rd_o / safe_wr_o", "bank_timer.sv:135-136",
        "safe_rd = safe_wr = r_row_valid && (r_rcd == 0) && !r_ap_pending",
        ["row_valid", "rcd==0", "ap_pending"],
        lambda rv, rcd0, ap: rv and rcd0 and (not ap),
        "Single 1-cell at (1,1,0). ap_pending=1 must kill columns — the row "
        "is committed to auto-close.")
    km.kmap(
        "safe_pre_o", "bank_timer.sv:138",
        "safe_pre = r_row_valid && (r_ras == 0) && (r_preblk == 0) && "
        "!r_ap_pending",
        ["row_valid", "ras==0", "preblk==0", "ap_pending"],
        lambda rv, ras0, pb0, ap: rv and ras0 and pb0 and (not ap),
        "Single 1-cell at (1,1,1,0): tRAS AND tRTP/tWR both elapsed, no "
        "auto-PRE in flight. preblk covers the read-to-PRE / write-recovery "
        "window the arbiter cannot see per-command.")
    km.kmap(
        "w_ap_fire (internal auto-precharge)", "bank_timer.sv:88",
        "w_ap_fire = r_ap_pending && (r_preblk == 0) && (r_ras == 0)",
        ["ap_pending", "preblk==0", "ras==0"],
        lambda ap, pb0, ras0: ap and pb0 and ras0,
        "Single 1-cell at (1,1,1). Fires exactly once: it clears ap_pending "
        "and row_valid and loads tRP the same edge.")
    km.kmap(
        "state_o (observability only)", "bank_timer.sv:144-147",
        "rv ? (rcd_nz ? ACTIVATING : ACTIVE) : (rp_nz ? PRECHARGING : IDLE)",
        ["row_valid", "rcd_nz", "rp_nz"],
        lambda rv, rcd, rp: ("ACTIVATING" if (rv and rcd) else
                             "ACTIVE" if rv else
                             "PRECHG" if rp else "IDLE"),
        "Pure decode of the flags; no downstream logic may depend on it.",
        values={})
    km.table(
        "row_valid / ap_pending next-state priority", "bank_timer.sv:115-127",
        ["set_act", "set_pre", "w_ap_fire", "set_rd||set_wr",
         "row_valid'", "ap_pending'"],
        [("1", "x", "x", "x", "1 (row=row_i)", "0"),
         ("0", "1", "x", "x", "0", "0"),
         ("0", "0", "1", "x", "0", "0"),
         ("0", "0", "0", "1", "unchanged", "set_ap_i"),
         ("0", "0", "0", "0", "unchanged", "unchanged")],
        note="Strict priority ACT > PRE > auto-PRE > column. The arbiter "
             "never issues ACT and PRE to one bank in one cycle "
             "(single-issue), so rows 1-2 cannot conflict in practice.")
    return ws


def build_refresh_sheet(wb):
    if "K-maps refresh+top" in wb.sheetnames:
        del wb["K-maps refresh+top"]
    ws = wb.create_sheet("K-maps refresh+top")
    for col, w in zip("ABCDEFGH", (26, 9, 9, 9, 9, 9, 9, 9)):
        ws.column_dimensions[col].width = w
    km = KmapWriter(ws)
    km.sheet_intro(
        "refresh_ctrl + scheduler top — K-maps (rtl/fub/refresh_ctrl.sv, "
        "rtl/macro/pumice_mem_cmd_scheduler.sv)",
        ["Flags: pend_nz=(r_pending>0), rem_nz=(r_burst_remaining>0), "
         "expired=(r_refi_cnt==0)."])

    km.kmap(
        "w_grant_accept", "refresh_ctrl.sv:79",
        "w_grant_accept = refresh_grant_i && (r_pending > 0)",
        ["grant", "pend_nz"],
        lambda g, p: g and p,
        "Single 1-cell at (1,1): a grant with nothing pending is swallowed "
        "(the accumulator never goes negative).")
    km.kmap(
        "refresh_drain_active", "refresh_ctrl.sv:126",
        "w_drain_active = (r_burst_remaining > 0) && (r_pending > 0)",
        ["rem_nz", "pend_nz"],
        lambda r, p: r and p,
        "Single 1-cell at (1,1); holding the arbiter in the refresh branch "
        "requires BOTH quota and owed refreshes.")
    km.table(
        "r_pending next-value", "refresh_ctrl.sv:98-108",
        ["enable && expired", "w_grant_accept", "next"],
        [("1", "0", "min(pending+1, 8)  (postpone cap)"),
         ("0", "1", "pending-1"),
         ("1", "1", "pending (net zero)"),
         ("0", "0", "pending")],
        note="refresh_req_o (registered) = pending>0. Saturation at 8 = a "
             "JEDEC retention violation looming — the arbiter must be "
             "serving REFs.")
    km.kmap(
        "busy_o", "pumice_mem_cmd_scheduler.sv",
        "busy = !init_done || refresh_req || w_cmd_rd_valid || "
        "(|w_bank_row_active[0])",
        ["init_done", "refresh_req", "cmd_rd_valid", "any_row_active"],
        lambda i, r, v, a: (not i) or r or v or a,
        "Zero ONLY at (1,0,0,0): init complete, no refresh owed, cmd FIFO "
        "empty, all rows closed. 15 of 16 cells are 1 — busy is the "
        "OR-reduce of everything in flight.")
    return ws


def refresh_contract_sheets(wb):
    """Update the two original sheets for the current RTL."""
    RESOLVED = ("RESOLVED 2026-07-22: the on-board device-word/shift story "
                "was host config + measurement (see #39 close-out) — read "
                "path clean at the bring-up tuple (wrlat=1/rden=6/rdly=7). "
                "Remaining integrity issue = runtime config axes, #42.")
    ws = wb["AXI to CAMs"]
    for row in ws.iter_rows(min_row=2):
        note = row[7].value or ""
        if "SUSPECT" in str(note):
            row[7].value = RESOLVED
            row[7].alignment = WRAP

    ws = wb["Scheduler both dirs"]
    # update the t_refi row's invariant for w_ref_safe + tRFC
    for row in ws.iter_rows(min_row=2):
        if row[1].value == "t_refi/refresh_burst":
            row[6].value = ("REF emitted every t_refi; REF only under "
                           "w_ref_safe (rows closed + nothing in flight/"
                           "guarded + prior tRFC elapsed); drain REFs "
                           "tRFC-spaced")
            row[6].alignment = WRAP
        if row[1].value == "cl_o/cwl_o/bl_o":
            row[5].value = ("CAS latency / CAS write latency / burst length "
                           "shadow from the MR writes. RTL reset = BL8 "
                           "(MR0.VAL 0x0433); the Nexys board runtime-"
                           "programs BL4 (0x0432).")
            row[5].alignment = WRAP
    # append the missing refresh/tRFC/checker rows
    new_rows = [
        ("Refresh iface", "refresh_req_i / refresh_drain_i", "1/1", "in",
         "refresh_ctrl",
         "Owed-refresh request (pending>0) and burst-drain hold.",
         "req high => arbiter priority-2 branch; columns/ACTs starve by "
         "design while high", ""),
        ("Refresh iface", "refresh_grant_o", "1", "out", "scheduler",
         "Pulsed when the REF command FIRES (w_fire_out && r_grant).",
         "one grant per issued REF; decrements pending + drain quota", ""),
        ("Timing cfg", "t_rfc_i", "16", "in", "CSR (TIMINGS_RFC_REFI.tRFC)",
         "Mission-mode REF->ACT/REF recovery, MC cycles (reset 16). Loaded "
         "into the arbiter's r_rfc_cnt on each fired REF.",
         "no ACT or REF while r_rfc_cnt != 0 (w_rfc_busy) — audited by the "
         "history checker's positional tRFC check",
         "added 2026-07-21: previously enforced by NOTHING in mission mode"),
        ("DV hooks", "CMD_HISTORY_EN / HIST_T_*", "param", "-", "generate",
         "In-scheduler command-history scoreboard (audit-only shift "
         "registers on the ISSUED stream; $fatal assertions, sim-only).",
         "default 0 = zero cost; -GCMD_HISTORY_EN=1 + --assert in DV",
         "moved into the scheduler 2026-07-22 (was a TB bind)"),
    ]
    base = ws.max_row + 1
    for i, vals in enumerate(new_rows):
        for j, v in enumerate(vals):
            c = ws.cell(base + i, 1 + j, v)
            c.alignment = WRAP


def main():
    wb = openpyxl.load_workbook(XLSX)
    refresh_contract_sheets(wb)
    build_arbiter_sheet(wb)
    build_bank_timer_sheet(wb)
    build_refresh_sheet(wb)
    wb.save(XLSX)
    print(f"updated {XLSX}")
    for name in wb.sheetnames:
        print(" ", name)


if __name__ == "__main__":
    main()
