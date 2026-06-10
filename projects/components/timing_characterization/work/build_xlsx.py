#!/usr/bin/env python3
"""Build asap7_stream_library_timing.xlsx from /tmp/charwork/asap7/timings.csv.

Three sheets:
  Library_Summary   - one row per library block; delay/fmax/area/gates across
                      all corner+freq runs.
  Per_Corner        - one row per (block, freq); columns for TT/FF/SS delay.
  Raw               - the full 99-row CSV unmodified.
"""
from __future__ import annotations
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

CSV = Path("/tmp/charwork/asap7/timings.csv")
OUT = Path("/mnt/data/github/RTLDesignSherpa/projects/components/timing_characterization/work/asap7_stream_library_timing.xlsx")

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
CENTER = Alignment(horizontal="center")

def read_rows():
    with CSV.open() as f:
        return list(csv.DictReader(f))

def style_header(ws, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

def autosize(ws):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        w = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col_letter].width = min(max(w + 2, 10), 50)

def build_summary(wb, rows):
    """One row per library block with delay/fmax/area aggregated across corners+freqs."""
    by_row: dict[str, list] = {}
    for r in rows:
        if r["status"] != "OK":
            continue
        by_row.setdefault(r["row"], []).append(r)

    ws = wb.active
    ws.title = "Library_Summary"
    cols = [
        "library_block",
        "actual_depth", "char_depth", "storage_bits",
        "delay_min_ps", "delay_max_ps", "fmax_max_MHz", "fmax_min_MHz",
        "gate_count", "ff_count", "area_um2", "notes",
    ]
    ws.append(cols)
    style_header(ws, len(cols))

    for label in sorted(by_row):
        rs = by_row[label]
        delays = [float(r["delay_ps"]) for r in rs]
        gate_count = int(rs[0]["gate_count"])
        ff_count   = int(rs[0]["ff_count"])
        area       = float(rs[0]["area_um2"])
        ws.append([
            label,
            int(rs[0]["actual_depth"]), int(rs[0]["char_depth"]),
            int(rs[0]["storage_bits"]),
            round(min(delays), 2), round(max(delays), 2),
            round(1_000_000 / min(delays), 0),
            round(1_000_000 / max(delays), 0),
            gate_count, ff_count, round(area, 1),
            rs[0]["notes"],
        ])
    autosize(ws)

def build_per_corner(wb, rows):
    """Pivot table: one row per (library_block, freq_MHz), columns TT/FF/SS delay."""
    by_key: dict[tuple, dict] = {}
    for r in rows:
        if r["status"] != "OK":
            continue
        k = (r["row"], int(r["freq_mhz"]))
        by_key.setdefault(k, {})[r["corner"]] = r

    ws = wb.create_sheet("Per_Corner")
    cols = [
        "library_block", "freq_MHz", "period_target_ps",
        "TT_delay_ps", "FF_delay_ps", "SS_delay_ps",
        "TT_slack_ps", "FF_slack_ps", "SS_slack_ps",
        "TT_fmax_MHz", "FF_fmax_MHz", "SS_fmax_MHz",
    ]
    ws.append(cols)
    style_header(ws, len(cols))
    for (label, freq) in sorted(by_key):
        d = by_key[(label, freq)]
        tt = d.get("TT", {}); ff = d.get("FF", {}); ss = d.get("SS", {})
        ws.append([
            label, freq, int(tt.get("period_target_ps") or ff.get("period_target_ps") or 0),
            float(tt["delay_ps"]) if tt else "",
            float(ff["delay_ps"]) if ff else "",
            float(ss["delay_ps"]) if ss else "",
            float(tt["slack_ps"]) if tt else "",
            float(ff["slack_ps"]) if ff else "",
            float(ss["slack_ps"]) if ss else "",
            float(tt["fmax_mhz"]) if tt else "",
            float(ff["fmax_mhz"]) if ff else "",
            float(ss["fmax_mhz"]) if ss else "",
        ])
    autosize(ws)

def build_raw(wb, rows):
    ws = wb.create_sheet("Raw")
    if not rows:
        return
    cols = list(rows[0].keys())
    ws.append(cols)
    style_header(ws, len(cols))
    for r in rows:
        ws.append([r.get(c, "") for c in cols])
    autosize(ws)

def main():
    rows = read_rows()
    wb = Workbook()
    build_summary(wb, rows)
    build_per_corner(wb, rows)
    build_raw(wb, rows)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    n_ok = sum(1 for r in rows if r["status"] == "OK")
    print(f"wrote {OUT}  ({n_ok}/{len(rows)} OK runs)")

if __name__ == "__main__":
    main()
