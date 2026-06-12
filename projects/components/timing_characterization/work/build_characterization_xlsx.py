#!/usr/bin/env python3
"""Build asap7_characterization.xlsx — the design-time prediction table.

Reads /tmp/charwork/asap7/timing_char_data.csv (two-point synth probes per
primitive per corner) and derives:

  - Tflop[corner]            = Tcq + Tsu, the flop-to-flop overhead
                               (derived from NAND chain, the cleanest probe)
  - per_level[gate, corner]  = average std-cell delay per logic level
                               for that primitive in that corner

Then for each (corner, freq) combination:

      max_levels[gate, corner, freq] = floor((period - Tflop) / per_level)

Writes a single sheet `characterization` with columns of corner_freq
permutations and rows that match the methodology table (Tcq+Tsu, max levels
per gate type, per-level gate delay).
"""
from __future__ import annotations
import csv
import math
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CSV_IN  = Path("/tmp/charwork/asap7/timing_char_data.csv")
OUT     = Path("/mnt/data/github/RTLDesignSherpa/projects/components/timing_characterization/work/asap7_characterization.xlsx")

CORNERS = ["TT", "FF", "SS"]
FREQS   = [500, 750, 1000]  # MHz

# Primitives used for the per-level lookup, in display order.
# - NAND/XOR/MUX: the most useful gates (clean two-point slope, simple to reason about)
# - CARRY/MULT/GRAY_CTR/QUEUE: scale cleanly with their probe param
# - MULT is the RECOMMENDED PROXY for "mixed-bag" combinational logic - a real
#   multiplier tree contains AND, XOR, full adder, and carry cells, so its per-
#   level delay is a sensible blended estimate for generic combinational stuff.
# - INV and CLK_DIV are listed in 'Reference single-point measurements' but
#   excluded from the per-level / max-levels tables because their lev count does
#   not change between the two probe sizes (INV: ABC collapses pairs;
#   CLK_DIV: critical path is intra-stage, not depth-scaled by NUM_STAGES).
GATE_PRIMS         = ["NAND", "XOR", "MUX", "CARRY", "MULT", "GRAY_CTR", "QUEUE"]
DEGENERATE_PRIMS   = ["INV", "CLK_DIV"]
TFLOP_PROBE        = "NAND"
MIXED_PROXY        = "MULT"  # highlight this row as the generic-mixed-logic proxy

# Filled in by build_sheet() so other sheets can pin to the design-clock cells.
DESIGN_CLOCK_ROW   = None

# ---------------------------------------------------------------------------- #
HDR_FONT = Font(bold=True, color="000000")
HDR_FILL = PatternFill("solid", fgColor="D9E1F2")
SECTION_FONT = Font(bold=True)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
NUM_FONT = Font()
CENTER = Alignment(horizontal="center")
THIN = Side(border_style="thin", color="A0A0A0")
BOX = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def read_probes() -> dict:
    """Return {(prim, corner): [(value, delay_ps, lev), ...]}."""
    probes: dict = {}
    with CSV_IN.open() as f:
        for r in csv.DictReader(f):
            if r["status"] != "OK":
                continue
            k = (r["primitive"], r["corner"])
            probes.setdefault(k, []).append(
                (int(r["value"]), float(r["delay_ps"]), int(r["lev"])))
    return probes


def derive(probes: dict) -> tuple[dict, dict, dict]:
    """Return (per_level, tflop, single_point_d) keyed by (gate, corner)/corner."""
    per_level: dict = {}
    tflop: dict = {}
    single_pt: dict = {}  # smallest-size delay for degenerate-slope prims
    all_prims = GATE_PRIMS + DEGENERATE_PRIMS
    for prim in all_prims:
        for corner in CORNERS:
            ps = probes.get((prim, corner))
            if not ps:
                continue
            ps_sorted = sorted(ps, key=lambda x: x[0])
            (_v1, d1, lev1) = ps_sorted[0]
            single_pt[(prim, corner)] = (d1, lev1)
            if len(ps) < 2:
                continue
            (_v2, d2, lev2) = ps_sorted[-1]
            if lev2 == lev1:
                continue  # degenerate slope (INV pairs, CLK_DIV intra-stage)
            per_lv = (d2 - d1) / (lev2 - lev1)
            per_level[(prim, corner)] = per_lv
            if prim == TFLOP_PROBE:
                tflop[corner] = d1 - lev1 * per_lv
    return per_level, tflop, single_pt


def build_sheet(wb: Workbook, per_level: dict, tflop: dict, single_pt: dict):
    ws = wb.active
    ws.title = "characterization"

    # Local style fills used by the design-clock cell.
    PARAM_FILL = PatternFill("solid", fgColor="FFF2CC")
    FORMULA_FILL = PatternFill("solid", fgColor="E7E6E6")

    # Column layout: row label | TT_500 | TT_750 | TT_1000 | FF_500 | ... | SS_1000
    col_headers = ["metric"] + [f"{c}_{f}MHz" for c in CORNERS for f in FREQS]
    ws.append(col_headers)
    for c in range(1, len(col_headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BOX

    # Reference: period per freq column
    row = ["Period (ps)"]
    for corner in CORNERS:
        for f in FREQS:
            row.append(round(1_000_000 / f, 1))
    ws.append(row)

    # Tflop per corner — same value across freq columns within a corner
    row = ["Clock-to-Out (Tcq + Tsu, ps)"]
    for corner in CORNERS:
        v = tflop.get(corner, 0)
        for _ in FREQS:
            row.append(round(v, 2))
    ws.append(row)

    # Blank visual separator
    ws.append([""] * len(col_headers))

    # Per-gate-type max-level rows
    section_header_row(ws, "Max combinational levels per cycle", len(col_headers))
    for gate in GATE_PRIMS:
        label = f"Max {gate} levels"
        if gate == MIXED_PROXY:
            label += "  (PROXY for mixed-bag logic)"
        row = [label]
        for corner in CORNERS:
            per_lv = per_level.get((gate, corner), 0)
            tf = tflop.get(corner, 0)
            for f in FREQS:
                period = 1_000_000 / f
                budget = period - tf
                if per_lv > 0 and budget > 0:
                    row.append(math.floor(budget / per_lv))
                else:
                    row.append("")
        ws.append(row)
        # Highlight the MULT row to visually call it out as the mixed-bag proxy
        if gate == MIXED_PROXY:
            for c in range(1, len(col_headers) + 1):
                ws.cell(row=ws.max_row, column=c).fill = PatternFill(
                    "solid", fgColor="FFF2CC")

    ws.append([""] * len(col_headers))

    # Per-level delay table (constants per corner)
    section_header_row(ws, "Per-level gate delay (ps)  - freq-independent", len(col_headers))
    for gate in GATE_PRIMS:
        label = f"per {gate} delay (ps)"
        if gate == MIXED_PROXY:
            label += "  (PROXY)"
        row = [label]
        for corner in CORNERS:
            v = per_level.get((gate, corner), 0)
            for _ in FREQS:
                row.append(round(v, 2))
        ws.append(row)
        if gate == MIXED_PROXY:
            for c in range(1, len(col_headers) + 1):
                ws.cell(row=ws.max_row, column=c).fill = PatternFill(
                    "solid", fgColor="FFF2CC")

    # Reference single-point measurements for degenerate-slope primitives
    ws.append([""] * len(col_headers))
    section_header_row(
        ws,
        "Reference single-point delays (ps) - degenerate slope, no per-level number",
        len(col_headers))
    for gate in DEGENERATE_PRIMS:
        row = [f"{gate} flop-to-flop delay (ps)"]
        for corner in CORNERS:
            d, _lev = single_pt.get((gate, corner), (0, 0))
            for _ in FREQS:
                row.append(round(d, 2))
        ws.append(row)

    # ---------------------------------------------------------------------- #
    # Design clock: one cell defines THE target frequency for the design.
    # The sheet's 9 corner_freq columns above are a SWEEP - this section
    # picks the single design target that the other sheets reference.
    # ---------------------------------------------------------------------- #
    ws.append([""] * len(col_headers))
    section_header_row(ws, "Design clock  -  single configurable target",
                       len(col_headers))
    # Editable frequency cell + derived period.  Place them in cols A..C so
    # they're easy to spot.  Other sheets can reference these directly.
    r = ws.max_row + 1
    ws.cell(row=r, column=1, value="Target clock frequency (MHz):")
    c = ws.cell(row=r, column=2, value=1000); c.fill = PARAM_FILL; c.alignment = CENTER
    ws.cell(row=r, column=3, value="Target clock period (ps):")
    c = ws.cell(row=r, column=4, value=f"=1000000/B{r}"); c.fill = FORMULA_FILL; c.alignment = CENTER
    # Stash row number so other sheets can reference the design clock cells.
    # Period cell is at characterization!$D${design_clock_row}.
    global DESIGN_CLOCK_ROW
    DESIGN_CLOCK_ROW = r

    # ---------------------------------------------------------------------- #
    # STA boundary timing budget defaults (input delay 60%, output delay 40%).
    # ---------------------------------------------------------------------- #
    ws.append([""] * len(col_headers))
    section_header_row(ws,
        "STA boundary timing budget defaults  -  input 60% / output 40% of design period",
        len(col_headers))
    r = ws.max_row + 1
    ws.cell(row=r, column=1, value="Input delay (60% of design period, ps):")
    c = ws.cell(row=r, column=2, value=f"=D{DESIGN_CLOCK_ROW}*0.6")
    c.fill = FORMULA_FILL; c.alignment = CENTER
    ws.cell(row=r + 1, column=1, value="Output delay (40% of design period, ps):")
    c = ws.cell(row=r + 1, column=2, value=f"=D{DESIGN_CLOCK_ROW}*0.4")
    c.fill = FORMULA_FILL; c.alignment = CENTER
    INPUT_DELAY_ROW = r
    OUTPUT_DELAY_ROW = r + 1
    # Stash the row numbers for later cross-sheet references (only used inside
    # this function; other sheets pin to characterization!$B${row}).
    ws._design_clock_period_ref = f"characterization!$D${DESIGN_CLOCK_ROW}"

    # ---------------------------------------------------------------------- #
    # Wire delay per mm by metal layer.  ASAP7 NLDM has no wire_load table, so
    # ABC reports WireLoad = "none" on every gate-delay run above.  These are
    # 7nm industry rules-of-thumb for *buffered* routes:
    #   M2/M3   intermediate, used for short FUB-internal nets
    #   M4/M5   semi-global, used for FUB-to-FUB inside a partition
    #   M6/M7   global, used for partition-to-partition
    #   M8/M9   thick, clock/power/bus
    # Corner scaling: FF ~ 0.80 x TT, SS ~ 1.30 x TT (R rises with temp).
    # Numbers are freq-independent (distance-based), repeated across freq cols
    # to keep the sheet shape uniform.
    # ---------------------------------------------------------------------- #
    WIRE_PS_PER_MM_TT = {
        "M2/M3 (local / intermediate)": 120.0,
        "M4/M5 (semi-global)":          65.0,
        "M6/M7 (global)":               40.0,
        "M8/M9 (thick)":                22.0,
    }
    CORNER_WIRE_SCALE = {"TT": 1.0, "FF": 0.8, "SS": 1.3}

    ws.append([""] * len(col_headers))
    section_header_row(
        ws,
        "Wire delay per mm by metal layer (buffered route, ps/mm) - "
        "cross-partition route estimate",
        len(col_headers))
    for layer, tt_val in WIRE_PS_PER_MM_TT.items():
        row = [layer]
        for corner in CORNERS:
            v = tt_val * CORNER_WIRE_SCALE[corner]
            for _ in FREQS:
                row.append(round(v, 1))
        ws.append(row)

    # Notes
    ws.append([""] * len(col_headers))
    ws.append(["Notes"])
    ws.cell(row=ws.max_row, column=1).font = SECTION_FONT
    notes = [
        "Probe corner_freq columns are independent measurements (see Per-level "
        "table above).",
        f"Tflop = Tcq + Tsu derived from {TFLOP_PROBE} chain two-point probe "
        "(LEVELS={3,6}).",
        "max_levels = floor((period_ps - Tflop) / per_level_delay).",
        f"** When the critical path is a mixed bag of cells (typical for "
        f"datapath / FSM next-state / arbitration logic), use the {MIXED_PROXY} "
        f"per-level number. It bakes in AND / XOR / full-adder / carry mix.",
        "INV: ABC's strash collapses inverter pairs across the chain, so lev "
        "doesn't change between probe sizes - per-level slope is degenerate. "
        "Single-point delay is reported in the reference table.",
        "CLK_DIV: critical path is intra-stage (counter + pickoff), so NUM_STAGES "
        "doesn't move lev. Single-point delay is reported in the reference table.",
        "Wire delays are 7nm industry rule-of-thumb for buffered routes; the "
        "ASAP7 NLDM ships no wire_load table so ABC measured gate delay only. "
        "For partition-to-partition output budgeting, look up the destination "
        "metal layer and multiply by route length (mm).",
        "Raw probe data is in work/timing_char_data.csv. Reproducer is "
        "work/timing_char_sweep.py.",
    ]
    for n in notes:
        ws.append([n])

    # Column widths
    ws.column_dimensions["A"].width = 38
    for c in range(2, len(col_headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12

    # Style data cells
    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            if cell.column > 1 and isinstance(cell.value, (int, float)):
                cell.alignment = CENTER


def section_header_row(ws, text, n_cols):
    ws.append([text] + [""] * (n_cols - 1))
    r = ws.max_row
    ws.cell(row=r, column=1).font = SECTION_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL


def build_blocks_sheet(wb: Workbook):
    """Per-block structural estimator that scales from the characterization sheet.

    Layout:
      Header row (column titles)
      'SRAMs' section header
         - raw SRAM macro
         - gaxi_fifo_sync (REG=1, SRAM-backed)
      'Building blocks' section header
         - gaxi_skid_buffer, gaxi_fifo_sync (REG=0), arbiter_RR,
           axi4_master_rd, axi4_master_wr

    Cells B..G are editable params (highlighted yellow). flop/gate/data->D
    delay cells are formulas keyed off the characterization sheet's Tflop /
    per_NAND / per_MUX cells.

    Cell refs into 'characterization':
      Tflop per corner:  $D$3 (TT)  $G$3 (FF)  $J$3 (SS)
      per NAND delay:    $D$15 (TT) $G$15 (FF) $J$15 (SS)
      per MUX  delay:    $D$17 (TT) $G$17 (FF) $J$17 (SS)
    """
    ws = wb.create_sheet("building blocks")
    PARAM_FILL = PatternFill("solid", fgColor="FFF2CC")  # yellow = editable
    FORMULA_FILL = PatternFill("solid", fgColor="E7E6E6")  # grey = computed
    # Section banners: black bold text on a high-contrast amber fill, so
    # the heading is readable even if a viewer fails to render the fill.
    BIG_HDR_FILL = PatternFill("solid", fgColor="FFC000")
    BIG_HDR_FONT = Font(bold=True, size=14, color="000000")

    header = [
        "block",                                       # A
        "P1", "P2", "P3", "P4", "P5", "P6",            # B-G editable params
        "params",                                      # H label string
        "flop_count",                                  # I
        "NAND_eq_gates",                               # J
        # ---- input-side: data -> first capturing flop's D ----
        "in MUX lv", "in NAND lv",                     # K, L
        "data->D TT (ps)", "data->D FF (ps)", "data->D SS (ps)",  # M, N, O
        # ---- output-side: internal flop Q -> module output port ----
        "out MUX lv", "out NAND lv",                   # P, Q
        "flop->out TT (ps)", "flop->out FF (ps)", "flop->out SS (ps)",  # R, S, T
        # ---- fmax = worst of the two internal halves ----
        "fmax TT (MHz)", "fmax FF (MHz)", "fmax SS (MHz)",  # U, V, W
        # ---- chain accumulator (per-corner) ----
        "delay_in (ps)",                                            # X
        "delay_out TT", "delay_out FF", "delay_out SS",             # Y, Z, AA
        "",                                                         # AB blank gap
        "slack TT (ps)", "slack FF (ps)", "slack SS (ps)",          # AC, AD, AE
        "notes",                                       # AF
    ]
    ws.append(header)
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BOX

    NCOL = len(header)

    # ---------- Config row 2: target period + 60% input / 40% output ---------- #
    ws.cell(row=2, column=1, value="CONFIG -> target period (ps):").font = SECTION_FONT
    c = ws.cell(row=2, column=2, value=1000); c.fill = PARAM_FILL; c.alignment = CENTER
    ws.cell(row=2, column=3, value="input delay 60% (ps):").font = SECTION_FONT
    c = ws.cell(row=2, column=4, value="=B2*0.6"); c.fill = FORMULA_FILL; c.alignment = CENTER
    ws.cell(row=2, column=5, value="output delay 40% (ps):").font = SECTION_FONT
    c = ws.cell(row=2, column=6, value="=B2*0.4"); c.fill = FORMULA_FILL; c.alignment = CENTER
    ws.cell(row=2, column=7, value=(
        "  delay_in defaults to 0.6*period (STA input arrival). Chain by "
        "=Y_prev. slack reserves 0.4*period for downstream output delay."))

    def emit_section_header(label: str):
        r = ws.max_row + 1
        ws.append([label] + [""] * (NCOL - 1))
        ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=NCOL)
        cell = ws.cell(row=r, column=1)
        cell.font = BIG_HDR_FONT
        cell.fill = BIG_HDR_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 22

    # ----------------------------------------------------------------------- #
    # Block definitions
    #   params: tuple of (label, base_value) — written to B..G in order
    #   flops:  Excel formula using {R} as the current row index for self-refs
    #   nand:   Excel formula for NAND-equivalent combinational gates
    #   mux_lv: Excel formula or literal int — MUX levels on data->D path
    #   nand_lv: Excel formula or literal int — NAND levels on data->D path
    #   notes:  string
    # ----------------------------------------------------------------------- #
    # SRAM section: storage primitives and their BRAM-backed FIFO wrapper.
    # For SRAM-backed designs the W*D data goes to a BRAM/SRAM macro - it is
    # NOT in the flop count.  The flop_count formula on these rows only counts
    # control flops (ptrs + count).
    SRAM_BLOCKS = [
        dict(
            name="SRAM (raw macro)",
            params=[("W", 512), ("D", 512)],
            flops=0,                                # storage is in the macro
            nand="=4*B{R} + 8*CEILING(LOG(C{R},2),1)",  # IO mux + addr decode
            in_mux=0, in_nand=1,                    # write addr mux is shallow
            out_mux=0, out_nand=1,                  # synchronous read - registered
            notes="SRAM macro (single-port, synchronous read). Storage = W*D "
                  "bits goes to the macro, not flop count. Read latency = 1 "
                  "cycle (data appears next clock). Replace this row's per-bit "
                  "delay with vendor SRAM datasheet for accurate timing.",
        ),
        dict(
            name="gaxi_fifo_sync (REG=1, SRAM-backed)",
            params=[("W", 512), ("D", 512)],
            # Only control flops: 2*log2(D) for wr/rd ptrs + log2(D)+1 for count.
            # Storage (W*D) lives in the BRAM/SRAM macro - not in flop count.
            flops="=3*CEILING(LOG(C{R},2),1) + 1",
            nand="=B{R}*4 + CEILING(LOG(C{R},2),1)*8 + 20",
            in_mux=0, in_nand=1,                    # write side: small ptr mux
            out_mux=0, out_nand=1,                  # registered out hides read mux
            notes="REGISTERED=1 + MEM_STYLE=BRAM: storage is in an SRAM/BRAM "
                  "macro, not flops. Used only by SRAM controller units in "
                  "STREAM/RAPIDS. Storage bits = W*D. fmax is limited by "
                  "the SRAM macro clock-to-Q, not by std-cell timing.",
        ),
    ]

    BUILDING_BLOCKS = [
        dict(
            name="gaxi_skid_buffer",
            params=[("W", 1), ("D", 2)],
            flops="=B{R}*C{R} + C{R} + 2",
            nand="=B{R}*C{R}*4 + 10",
            # Skid buffers: output is the registered head flop Q directly.
            # NO output mux - the head data flop is what drives the output port.
            in_mux=1,  in_nand=1,                   # bypass MUX2 on input
            out_mux=0, out_nand=0,                  # direct Q to output port
            notes="Skid buffer outputs come straight from the head flop Q - "
                  "no output muxing. Input side has the bypass MUX2. "
                  "Flops = W*D storage + D valid + 2 control.",
        ),
        dict(
            name="gaxi_fifo_sync (REG=0)",
            params=[("W", 1), ("D", 2)],
            flops="=B{R}*C{R} + 3*CEILING(LOG(C{R},2),1) + 1",
            nand="=B{R}*C{R}*6 + CEILING(LOG(C{R},2),1)*8 + 20",
            # Read mux is on the OUTPUT side - log2(D) levels of MUX
            in_mux=1, in_nand=2,                                   # write decode + full cmp
            out_mux="=CEILING(LOG(C{R},2),1)", out_nand=0,         # read mux tree
            notes="Output critical path is the READ mux tree, depth=log2(D). "
                  "Flops = W*D storage + 3*log2(D) ptrs/count + 1 state. "
                  "Default FIFO in STREAM/RAPIDS (non-SRAM-backed).",
        ),
        dict(
            name="arbiter_round_robin",
            params=[("N", 2)],
            flops="=B{R} + 2",
            nand="=B{R}*8",
            # Priority encoder is output-side: req -> encode -> gnt
            in_mux=0, in_nand=1,                                   # request register
            out_mux="=CEILING(LOG(B{R},2),1)", out_nand=1,         # encoder tree
            notes="Priority encoder is log2(N) MUX levels on the OUTPUT side "
                  "(req -> encode -> gnt). Flops = N (last-grant tracker) + 2.",
        ),
        dict(
            name="axi4_master_rd",
            params=[("AW", 32), ("DW", 32), ("IW", 4), ("UW", 1),
                    ("SKID_AR", 2), ("SKID_R", 2)],
            # AR slot width = AW+IW+UW+29 (per axi4_master_rd ARSize formula);
            # R slot width  = IW+DW+UW+3 (per RSize).
            flops="=F{R}*(B{R}+D{R}+E{R}+29) + G{R}*(D{R}+C{R}+E{R}+3)",
            nand="=4*(F{R}*(B{R}+D{R}+E{R}+29) + G{R}*(D{R}+C{R}+E{R}+3))",
            in_mux=1, in_nand=1,                    # input skid bypass MUX
            out_mux=1, out_nand=1,                  # output side skid select MUX
            notes="Two skid buffers in series (AR + R). Input bypass MUX2 + "
                  "output select MUX2. Flops scale with SKID_AR*ARw + SKID_R*Rw.",
        ),
        dict(
            name="axi4_master_wr",
            params=[("AW", 32), ("DW", 32), ("IW", 4), ("UW", 1),
                    ("SKID_AW", 2), ("SKID_W", 2)],
            # AW slot width = AW+IW+UW+29; W slot width = DW + DW/8 + UW + 1;
            # B slot width = IW + UW + 2.  Assume SKID_B=2 for the constant.
            flops="=F{R}*(B{R}+D{R}+E{R}+29) + G{R}*(C{R}+C{R}/8+E{R}+1) + 2*(D{R}+E{R}+2)",
            nand="=4*(F{R}*(B{R}+D{R}+E{R}+29) + G{R}*(C{R}+C{R}/8+E{R}+1) + 2*(D{R}+E{R}+2))",
            in_mux=1, in_nand=1,                    # input skid bypass MUX
            out_mux=1, out_nand=1,                  # output side skid select MUX
            notes="Three skid buffers (AW + W + B); SKID_B fixed at 2 in this "
                  "row. Input bypass + output select MUXes; both are 1 level.",
        ),
    ]

    # Characterization cell refs (per_NAND / per_MUX / Tflop, by corner).
    # NOTE: After expanding to 7 gate types, per_NAND moved to row 15 and
    # per_MUX moved to row 17 on the characterization sheet.
    REFS = {
        "TT": {"nand": "$D$15", "mux": "$D$17", "tflop": "$D$3"},
        "FF": {"nand": "$G$15", "mux": "$G$17", "tflop": "$G$3"},
        "SS": {"nand": "$J$15", "mux": "$J$17", "tflop": "$J$3"},
    }

    def write_block(b):
        r = ws.max_row + 1  # current target row
        params = b["params"]
        # Block name (col A)
        ws.cell(row=r, column=1, value=b["name"]).font = SECTION_FONT
        # Params P1..P6 (cols B..G)
        for i, (_lbl, v) in enumerate(params):
            cell = ws.cell(row=r, column=2 + i, value=v)
            cell.fill = PARAM_FILL
            cell.alignment = CENTER
        # Param-label string (col H)
        ws.cell(row=r, column=8,
                value=", ".join(f"P{i+1}={lbl}" for i, (lbl, _) in enumerate(params)))
        # I flops, J nand_eq — allow literal int (SRAM macro flops=0)
        flops_v = b["flops"]
        if isinstance(flops_v, str):
            flops_v = flops_v.format(R=r)
        ws.cell(row=r, column=9, value=flops_v)
        ws.cell(row=r, column=10, value=b["nand"].format(R=r))
        # K, L: input-side MUX/NAND levels
        for col, key in ((11, "in_mux"), (12, "in_nand")):
            v = b[key]
            if isinstance(v, str):
                v = v.format(R=r)
            ws.cell(row=r, column=col, value=v)
        # M, N, O: data->D combo delay per corner (input side)
        for off, corner in enumerate(("TT", "FF", "SS")):
            ref = REFS[corner]
            f = (f"=K{r}*characterization!{ref['mux']} "
                 f"+ L{r}*characterization!{ref['nand']}")
            cell = ws.cell(row=r, column=13 + off, value=f)
            cell.alignment = CENTER
            cell.fill = FORMULA_FILL
        # P, Q: output-side MUX/NAND levels
        for col, key in ((16, "out_mux"), (17, "out_nand")):
            v = b[key]
            if isinstance(v, str):
                v = v.format(R=r)
            ws.cell(row=r, column=col, value=v)
        # R, S, T: flop->out combo delay per corner (output side)
        for off, corner in enumerate(("TT", "FF", "SS")):
            ref = REFS[corner]
            f = (f"=P{r}*characterization!{ref['mux']} "
                 f"+ Q{r}*characterization!{ref['nand']}")
            cell = ws.cell(row=r, column=18 + off, value=f)
            cell.alignment = CENTER
            cell.fill = FORMULA_FILL
        # U, V, W: fmax per corner — worst of (data->D, flop->out) + Tflop
        for off, corner in enumerate(("TT", "FF", "SS")):
            ref = REFS[corner]
            in_col  = chr(ord("M") + off)  # M, N, O = data->D
            out_col = chr(ord("R") + off)  # R, S, T = flop->out
            f = (f"=ROUND(1000000/(MAX({in_col}{r},{out_col}{r}) "
                 f"+ characterization!{ref['tflop']}), 0)")
            cell = ws.cell(row=r, column=21 + off, value=f)
            cell.alignment = CENTER
            cell.fill = FORMULA_FILL
        # X: delay_in - default = 0.6 * target period (STA input arrival 60%).
        c = ws.cell(row=r, column=24, value="=$D$2"); c.fill = PARAM_FILL; c.alignment = CENTER
        # Y, Z, AA: delay_out TT/FF/SS = delay_in + (data->D + flop->out) per corner
        for off, in_col, out_col in (
            (0, "M", "R"),   # TT
            (1, "N", "S"),   # FF
            (2, "O", "T"),   # SS
        ):
            c = ws.cell(row=r, column=25 + off,
                        value=f"=X{r} + {in_col}{r} + {out_col}{r}")
            c.fill = FORMULA_FILL; c.alignment = CENTER
        # AB (col 28) intentionally blank as separator
        # AC, AD, AE: slack TT/FF/SS = period - delay_out - output_delay(40%)
        for off in range(3):
            out_col = get_column_letter(25 + off)  # Y, Z, AA (handles multi-letter)
            c = ws.cell(row=r, column=29 + off,
                        value=f"=$B$2 - {out_col}{r} - $F$2")
            c.fill = FORMULA_FILL; c.alignment = CENTER
        # AF: notes
        ws.cell(row=r, column=32, value=b["notes"])

    # Emit SRAMs section
    emit_section_header("SRAMs")
    for b in SRAM_BLOCKS:
        write_block(b)

    # Emit Building Blocks section
    emit_section_header("Building Blocks")
    for b in BUILDING_BLOCKS:
        write_block(b)

    # Storage-size reference at the bottom. After adding the config row at 2,
    # the SRAM rows now live at rows 4 and 5 (banner=3, raw=4, REG=1=5).
    r = ws.max_row + 2
    ws.cell(row=r, column=1, value="SRAM storage size reference").font = SECTION_FONT
    ws.cell(row=r + 1, column=1, value="SRAM raw macro: storage bits = W * D = ")
    ws.cell(row=r + 1, column=4, value="=B4*C4")
    ws.cell(row=r + 1, column=5, value="bits =")
    ws.cell(row=r + 1, column=6, value="=B4*C4/8/1024")
    ws.cell(row=r + 1, column=7, value="KB")
    ws.cell(row=r + 2, column=1, value="FIFO REG=1 SRAM-backed: storage bits = W * D = ")
    ws.cell(row=r + 2, column=4, value="=B5*C5")
    ws.cell(row=r + 2, column=5, value="bits =")
    ws.cell(row=r + 2, column=6, value="=B5*C5/8/1024")
    ws.cell(row=r + 2, column=7, value="KB")

    # Column widths
    widths = {1: 32, 8: 38, 9: 11, 10: 13,
              11: 8, 12: 9,                          # K, L  in mux/nand lv
              13: 15, 14: 15, 15: 15,                # M, N, O data->D
              16: 8, 17: 9,                          # P, Q  out mux/nand lv
              18: 15, 19: 15, 20: 15,                # R, S, T flop->out
              21: 13, 22: 13, 23: 13,                # U, V, W fmax
              24: 14,                                # X delay_in
              25: 14, 26: 14, 27: 14,                # Y, Z, AA delay_out TT/FF/SS
              28: 3,                                 # AB blank gap
              29: 13, 30: 13, 31: 13,                # AC, AD, AE slack TT/FF/SS
              32: 60}                                # AF notes
    for c in range(2, 8):
        widths.setdefault(c, 9)
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # Legend
    legend_row = ws.max_row + 2
    ws.cell(row=legend_row, column=1, value="Cell coding").font = SECTION_FONT
    ws.cell(row=legend_row + 1, column=1, value="Yellow = editable param").fill = PARAM_FILL
    ws.cell(row=legend_row + 2, column=1, value="Grey = derived formula").fill = FORMULA_FILL
    ws.cell(row=legend_row + 4, column=1, value=(
        "data->D = critical combinational delay on the path that loads the "
        "downstream flop. Compare against (period - Tflop) to size for a freq."
    ))
    ws.cell(row=legend_row + 5, column=1, value=(
        "Estimates are first-order: structural flop/gate counts + characterization "
        "per-level delays. Treat them as design-time SWAGs, not post-synth truth."
    ))
    ws.cell(row=legend_row + 6, column=1, value=(
        "REG=1 SRAM-backed FIFO is only used by SRAM controller units in "
        "STREAM/RAPIDS (sram_controller_unit, src/snk_sram_controller_unit_beats). "
        "Everywhere else uses REG=0 - that's why the Building Blocks section "
        "lists only the REG=0 variant."
    ))


def build_stream_sheet(wb: Workbook):
    """Per-FUB port summary for every STREAM FUB.

    Each FUB gets its own banner row, then one row per significant port with:
      - direction (in/out)
      - building block crossed (if any)
      - local NAND/MUX levels between port and nearest flop in this FUB
      - source/sink flop or external port
      - 'places it goes' = downstream consumer / upstream producer in the
        wider STREAM hierarchy (stream_core / scheduler_group / etc.)

    Local NAND depth/count is hand-walked from the RTL. Levels reflect cost
    *inside this FUB only*; building-block costs are on the 'building blocks'
    sheet.

    Refs (after 9-prim expansion):
      per_NAND: characterization!$D$15 (TT)  $G$15 (FF)  $J$15 (SS)
      per_MUX:  characterization!$D$17 (TT)  $G$17 (FF)  $J$17 (SS)
    """
    ws = wb.create_sheet("stream")

    PARAM_FILL = PatternFill("solid", fgColor="FFF2CC")
    FORMULA_FILL = PatternFill("solid", fgColor="E7E6E6")
    BIG_HDR_FILL = PatternFill("solid", fgColor="FFC000")
    BIG_HDR_FONT = Font(bold=True, size=14, color="000000")

    header = [
        "FUB", "port", "clock", "dir", "building block crossed",        # A-E
        "MUX lv", "NAND lv",                                            # F, G
        "source/sink flop or BB port",                                  # H
        "local delay TT (ps)", "local delay FF (ps)", "local delay SS (ps)",  # I, J, K
        "delay_in (ps)",                                                # L
        "delay_out TT", "delay_out FF", "delay_out SS",                 # M, N, O
        "",                                                             # P blank gap
        "slack TT (ps)", "slack FF (ps)", "slack SS (ps)",              # Q, R, S
        "places it goes (consumer / producer)",                         # T
    ]
    ws.append(header)
    NCOL = len(header)
    for c in range(1, NCOL + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BOX

    # Config row 2: target period + 60/40 boundary defaults
    ws.cell(row=2, column=1, value="CONFIG -> target period (ps):").font = SECTION_FONT
    c = ws.cell(row=2, column=2, value=1000); c.fill = PARAM_FILL; c.alignment = CENTER
    ws.cell(row=2, column=3, value="input 60% (ps):").font = SECTION_FONT
    c = ws.cell(row=2, column=4, value="=B2*0.6"); c.fill = FORMULA_FILL; c.alignment = CENTER
    ws.cell(row=2, column=5, value="output 40% (ps):").font = SECTION_FONT
    c = ws.cell(row=2, column=6, value="=B2*0.4"); c.fill = FORMULA_FILL; c.alignment = CENTER
    ws.cell(row=2, column=7, value=(
        "delay_in defaults to 0.6*period; chain by =M_prev. slack reserves "
        "0.4*period for downstream output delay."))

    def emit_fub_banner(label: str):
        r = ws.max_row + 1
        ws.append([label] + [""] * (NCOL - 1))
        ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=NCOL)
        cell = ws.cell(row=r, column=1)
        cell.font = BIG_HDR_FONT
        cell.fill = BIG_HDR_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 22

    REFS = {"TT": {"nand": "$D$15", "mux": "$D$17"},
            "FF": {"nand": "$G$15", "mux": "$G$17"},
            "SS": {"nand": "$J$15", "mux": "$J$17"}}

    def emit_row(fub, port, direction, bb, mux_lv, nand_lv, flop, goes, clock="clk"):
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=fub)
        ws.cell(row=r, column=2, value=port)
        ws.cell(row=r, column=3, value=clock).alignment = CENTER
        ws.cell(row=r, column=4, value=direction).alignment = CENTER
        ws.cell(row=r, column=5, value=bb)
        c = ws.cell(row=r, column=6, value=mux_lv); c.fill = PARAM_FILL; c.alignment = CENTER
        c = ws.cell(row=r, column=7, value=nand_lv); c.fill = PARAM_FILL; c.alignment = CENTER
        ws.cell(row=r, column=8, value=flop)
        # I, J, K: local delay TT/FF/SS  (mux_lv col=F, nand_lv col=G)
        for off, corner in enumerate(("TT", "FF", "SS")):
            f = (f"=F{r}*characterization!{REFS[corner]['mux']} "
                 f"+ G{r}*characterization!{REFS[corner]['nand']}")
            c = ws.cell(row=r, column=9 + off, value=f)
            c.fill = FORMULA_FILL; c.alignment = CENTER
        # L: delay_in (single cell, default 60% of period)
        c = ws.cell(row=r, column=12, value="=$D$2"); c.fill = PARAM_FILL; c.alignment = CENTER
        # M, N, O: delay_out TT/FF/SS = delay_in + local TT/FF/SS delay
        for off in range(3):
            local_col = chr(ord("I") + off)  # I, J, K
            c = ws.cell(row=r, column=13 + off, value=f"=L{r} + {local_col}{r}")
            c.fill = FORMULA_FILL; c.alignment = CENTER
        # P (col 16) intentionally blank as separator
        # Q, R, S: slack TT/FF/SS = period - delay_out - output_delay(40%)
        for off in range(3):
            out_col = chr(ord("M") + off)  # M, N, O
            c = ws.cell(row=r, column=17 + off,
                        value=f"=$B$2 - {out_col}{r} - $F$2")
            c.fill = FORMULA_FILL; c.alignment = CENTER
        # T: places it goes
        ws.cell(row=r, column=20, value=goes)

    # ----------------------------------------------------------------------- #
    # FUB definitions
    # port, dir, building block crossed, mux_lv, nand_lv, source/sink flop, places it goes
    # ----------------------------------------------------------------------- #
    FUBS = [
        ("stream_latency_bridge", [
            # local flop: r_drain_ip (the only one - rest is BB)
            ("s_valid",        "in",  "",                1, 1, "r_drain_ip (D)",
             "from u_channel_fifo.rd_valid in sram_controller_unit"),
            ("s_data[DW-1:0]", "in",  "gaxi_fifo_sync",  0, 0, "u_skid_buffer.wr_data",
             "from u_channel_fifo.rd_data"),
            ("r_drain_ip",     "local","",               0, 0, "u_skid_buffer.wr_valid (D)",
             "internal: only local flop; gates the skid push"),
            ("s_ready",        "out", "",                1, 2, "comb of u_skid_buffer.count + r_drain_ip",
             "to u_channel_fifo.rd_ready"),
            ("m_valid",        "out", "gaxi_fifo_sync",  0, 0, "u_skid_buffer.rd_valid (Q)",
             "to sram_controller_unit.axi_wr_sram_valid -> axi_write_engine"),
            ("m_data[DW-1:0]", "out", "gaxi_fifo_sync",  0, 0, "u_skid_buffer.rd_data (Q)",
             "to sram_controller_unit.axi_wr_sram_data -> axi_write_engine"),
            ("m_ready",        "in",  "gaxi_fifo_sync",  0, 0, "u_skid_buffer.rd_ready",
             "from axi_write_engine via sram_controller_unit"),
            ("occupancy[2:0]", "out", "gaxi_fifo_sync",  0, 0, "u_skid_buffer.count (Q)",
             "to scheduler (drain-IP backpressure)"),
            ("dbg_r_pending",  "out", "",                0, 0, "r_drain_ip (Q)",
             "debug observability"),
        ]),
        ("stream_alloc_ctrl", [
            # local flops: r_wr_ptr, r_rd_ptr, r_space_free, r_wr_full,
            # r_wr_almost_full, r_rd_empty, r_rd_almost_empty
            ("wr_valid",       "in",  "",                0, 1, "r_wr_ptr (D)",
             "from axi_read_engine (allocate request)"),
            ("wr_size[7:0]",   "in",  "",                0, 1, "r_wr_ptr adder input",
             "from axi_read_engine"),
            ("rd_valid",       "in",  "",                0, 1, "r_rd_ptr (D)",
             "from sram_controller_unit (release event)"),
            ("r_wr_ptr",       "local","",               0, 1, "r_space_free (D)",
             "internal: wr_ptr - rd_ptr = space_free"),
            ("r_rd_ptr",       "local","",               0, 1, "r_space_free (D)",
             "internal: subtractor into space_free"),
            ("r_space_free",   "local","",               0, 1, "r_wr_full (D)",
             "internal: cmp against DEPTH for full/almost_full flops"),
            ("space_free[AW:0]","out","",                0, 0, "r_space_free (Q)",
             "to axi_read_engine (sram_alloc_space_free)"),
            ("wr_full",        "out", "",                0, 0, "r_wr_full (Q)",
             "to axi_read_engine (back-pressure)"),
            ("wr_almost_full", "out", "",                0, 0, "r_wr_almost_full (Q)",
             "to axi_read_engine (threshold)"),
        ]),
        ("stream_drain_ctrl", [
            # local flops: r_wr_ptr, r_rd_ptr, r_data_available, r_rd_empty,
            # r_rd_almost_empty, r_wr_full
            ("wr_valid",       "in",  "",                0, 1, "r_wr_ptr (D)",
             "from u_channel_fifo handshake in sram_controller_unit"),
            ("rd_valid",       "in",  "",                0, 1, "r_rd_ptr (D)",
             "from axi_write_engine (drain reservation)"),
            ("rd_size[7:0]",   "in",  "",                0, 1, "r_rd_ptr adder input",
             "from axi_write_engine"),
            ("r_wr_ptr",       "local","",               0, 1, "r_data_available (D)",
             "internal: wr_ptr - rd_ptr = data_available"),
            ("r_rd_ptr",       "local","",               0, 1, "r_data_available (D)",
             "internal: subtractor into data_available"),
            ("r_data_available","local","",              0, 1, "r_rd_empty (D)",
             "internal: cmp 0 / cmp DEPTH for empty / full flops"),
            ("data_available[AW:0]","out","",            0, 0, "r_data_available (Q)",
             "to axi_write_engine"),
            ("rd_empty",       "out", "",                0, 0, "r_rd_empty (Q)",
             "to axi_write_engine (no data)"),
            ("wr_full",        "out", "",                0, 0, "r_wr_full (Q)",
             "to sram_controller_unit (back-pressure)"),
        ]),
        ("sram_controller_unit", [
            # local flops: r_axi_wr_sram_valid (registered output gate to write engine),
            # r_drain_ack (drain handshake), r_alloc_ack
            ("axi_rd_alloc_req",  "in",  "",             0, 1, "u_alloc_ctrl.wr_valid",
             "from axi_read_engine (reserve burst space)"),
            ("axi_rd_alloc_size[7:0]","in","",           0, 0, "u_alloc_ctrl.wr_size",
             "from axi_read_engine"),
            ("r_axi_wr_sram_valid","local","",           0, 0, "axi_wr_sram_valid (Q)",
             "registers latency_bridge.m_valid before driving the boundary"),
            ("axi_rd_alloc_space_free[AW:0]","out","",   0, 0, "u_alloc_ctrl.space_free (Q)",
             "to axi_read_engine"),
            ("axi_rd_sram_valid","in",  "gaxi_fifo_sync",0, 0, "u_channel_fifo.wr_valid",
             "from axi_read_engine (write data into SRAM FIFO)"),
            ("axi_rd_sram_data[DW-1:0]","in","gaxi_fifo_sync",0,0,"u_channel_fifo.wr_data",
             "from axi_read_engine"),
            ("axi_rd_sram_ready","out","gaxi_fifo_sync", 0, 0, "u_channel_fifo.wr_ready (Q)",
             "to axi_read_engine"),
            ("axi_wr_sram_valid","out","",               0, 0, "r_axi_wr_sram_valid (Q)",
             "to axi_write_engine"),
            ("axi_wr_sram_data[DW-1:0]","out","gaxi_fifo_sync",0,0,"u_latency_bridge.m_data (Q)",
             "to axi_write_engine"),
            ("axi_wr_sram_ready","in","",                0, 0, "u_latency_bridge.m_ready",
             "from axi_write_engine"),
            ("wr_drain_req",   "in",  "",                0, 1, "u_drain_ctrl.rd_valid",
             "from axi_write_engine"),
        ]),
        ("sram_controller", [
            # Mostly pure wrapping of N sram_controller_unit instances.
            # Local logic = channel-id mux/demux + N parallel SCU instances.
            ("ch_alloc_req[N-1:0]",  "in",  "",          1, 1, "u_unit[ch].axi_rd_alloc_req (D)",
             "from axi_read_engine (N-channel fan-out)"),
            ("ch_alloc_space_free[N-1:0][AW:0]","out","",1, 0, "u_unit[ch].axi_rd_alloc_space_free",
             "to axi_read_engine"),
            ("ch_rd_valid[N-1:0]","in", "",              1, 1, "u_unit[ch].axi_rd_sram_valid (D)",
             "from axi_read_engine"),
            ("ch_wr_valid[N-1:0]","out","",              1, 0, "u_unit[ch].axi_wr_sram_valid (Q)",
             "to axi_write_engine"),
            ("ch_wr_data[N-1:0][DW-1:0]","out","",       1, 0, "u_unit[ch].axi_wr_sram_data (Q)",
             "to axi_write_engine"),
        ]),
        ("perf_profiler", [
            # local flops: r_timestamp[63:0], r_channel_idle_prev[N-1:0],
            # r_event_pending, r_event_type, r_event_channel, r_fifo_wr_valid
            ("channel_idle[N-1:0]","in", "",             0, 1, "r_channel_idle_prev (D)",
             "from scheduler (per-ch idle state)"),
            ("cfg_enable",     "in",  "",                0, 1, "r_enable (D)",
             "from APB CSR"),
            ("cfg_clear",      "in",  "",                0, 1, "FIFO async-clear path",
             "from APB CSR"),
            ("r_timestamp[63:0]","local","",             0, 1, "r_event_payload (D)",
             "free-running counter; +1 every cycle when enabled"),
            ("r_channel_idle_prev","local","",           0, 1, "r_event_pending (D)",
             "XOR with channel_idle = edge detect"),
            ("r_event_pending","local","",               0, 1, "u_perf_fifo.wr_valid (D)",
             "1 cycle pulse when edge detected"),
            ("r_event_type",   "local","",               0, 0, "u_perf_fifo.wr_data[high] (D)",
             "rise / fall event code"),
            ("perf_fifo_rd",   "in",  "gaxi_fifo_sync",  0, 0, "u_perf_fifo.rd_valid",
             "from APB CSR pop strobe"),
            ("perf_fifo_data_low[31:0]","out","gaxi_fifo_sync",0,0,"u_perf_fifo.rd_data[31:0] (Q)",
             "to APB CSR (host read)"),
            ("perf_fifo_data_high[31:0]","out","gaxi_fifo_sync",0,0,"u_perf_fifo.rd_data[63:32] (Q)",
             "to APB CSR"),
            ("perf_fifo_empty","out","gaxi_fifo_sync",   0, 0, "u_perf_fifo.empty (Q)",
             "to APB CSR"),
            ("perf_fifo_count[15:0]","out","gaxi_fifo_sync",0,0,"u_perf_fifo.count (Q)",
             "to APB CSR"),
        ]),
        ("apbtodescr", [
            # local flops: r_apb_cmd_*, r_apb_rsp_*, r_ch_addr[N-1:0],
            # r_ch_valid[N-1:0], r_kickoff_hit, r_state
            ("apb_cmd_valid",  "in",  "",                0, 1, "r_apb_cmd_valid (D)",
             "from APB master"),
            ("apb_cmd_addr[ADDR_WIDTH-1:0]","in","",     0, 1, "r_state addr-decode (D)",
             "from APB master"),
            ("apb_cmd_wdata[DATA_WIDTH-1:0]","in","",    1, 1, "r_ch_addr[ch] (D)",
             "from APB master (descriptor pointer write)"),
            ("r_state",        "local","",               0, 1, "r_apb_rsp_valid (D)",
             "1-state FSM: ACCEPT -> RSP"),
            ("r_ch_addr[N-1:0]","local","",              0, 0, "desc_apb_addr (Q)",
             "per-channel registered descriptor pointer"),
            ("r_ch_valid[N-1:0]","local","",             0, 1, "desc_apb_valid (Q)",
             "per-channel kickoff valid; clears on desc_engine ack"),
            ("apb_cmd_ready",  "out", "",                0, 1, "comb (FIFO ready AND)",
             "to APB master"),
            ("apb_rsp_valid",  "out", "",                0, 0, "r_apb_rsp_valid (Q)",
             "to APB master"),
            ("desc_apb_valid[N-1:0]","out","",           0, 0, "r_ch_valid (Q)",
             "to descriptor_engine[N] (kick-off)"),
            ("desc_apb_addr[N-1:0][DA-1:0]","out","",    0, 0, "r_ch_addr (Q)",
             "to descriptor_engine[N] (initial descriptor pointer)"),
            ("desc_apb_ready[N-1:0]","in", "",           1, 1, "clears r_ch_valid[ch]",
             "from descriptor_engine[N]"),
        ]),
        ("descriptor_engine", [
            # local flops: r_state (FSM), r_desc_addr, r_descriptor (latch),
            # r_rdata_valid, r_burst_count, r_chain_valid
            ("desc_apb_valid", "in",  "",                0, 1, "r_state (D)",
             "from apbtodescr (kick-off)"),
            ("desc_apb_addr[DA-1:0]","in","",            0, 0, "r_desc_addr (D)",
             "from apbtodescr"),
            ("r_state",        "local","",               0, 1, "AR skid wr_valid (D)",
             "FSM IDLE -> FETCH -> WAIT_RDATA"),
            ("r_desc_addr",    "local","",               0, 0, "AR skid wr_data (D)",
             "drives ARADDR into the AR skid buffer"),
            ("r_descriptor[256:0]","local","",           0, 0, "i_descriptor_fifo.wr_data (D)",
             "captured descriptor data from R-channel"),
            ("r_chain_valid",  "local","",               0, 1, "r_desc_addr (D, next pointer)",
             "decodes next_descriptor_ptr -> re-arm FSM"),
            ("desc_axi_ar_*",  "out", "gaxi_skid_buffer",0, 0, "AR skid head Q",
             "to scheduler_group desc_axi master"),
            ("desc_axi_r_*",   "in",  "gaxi_skid_buffer",0, 0, "R skid input",
             "from scheduler_group desc_axi master"),
            ("descriptor_valid","out","gaxi_fifo_sync",  0, 0, "i_descriptor_fifo.rd_valid (Q)",
             "to scheduler (descriptor dispatch)"),
            ("descriptor_data[256-1:0]","out","gaxi_fifo_sync",0,0,"i_descriptor_fifo.rd_data (Q)",
             "to scheduler"),
            ("descriptor_ready","in","gaxi_fifo_sync",   0, 0, "i_descriptor_fifo.rd_ready",
             "from scheduler"),
        ]),
        ("scheduler", [
            # local flops: r_state (FSM), r_descriptor (latched), r_src_addr,
            # r_dst_addr, r_length, r_beats_remaining, r_rd_outstanding,
            # r_wr_outstanding, r_axi_rd_req, r_axi_wr_req, r_done
            ("descriptor_valid","in", "",                0, 1, "r_state (D)",
             "from descriptor_engine"),
            ("descriptor_data[256-1:0]","in","",         0, 1, "r_descriptor (D)",
             "from descriptor_engine: latch on accept"),
            ("r_descriptor",   "local","",               0, 0, "r_src_addr, r_dst_addr, r_length (D)",
             "field unpack: SRC[63:0], DST[127:64], LEN[159:128]"),
            ("r_state",        "local","",               0, 1, "r_axi_rd_req / r_axi_wr_req (D)",
             "FSM IDLE -> ISSUE_RD -> ISSUE_WR -> WAIT"),
            ("r_src_addr",     "local","",               0, 1, "r_axi_rd_req addr-incr (D)",
             "src_addr + (beats_issued * DW/8)"),
            ("r_beats_remaining","local","",             0, 1, "r_state DONE transition (D)",
             "decrement on each request; 0 -> COMPLETE"),
            ("r_rd_outstanding","local","",              0, 1, "r_axi_rd_req throttle (D)",
             "request flow-control counter"),
            ("axi_rd_req_valid","out","",                0, 0, "r_axi_rd_req (Q)",
             "to axi_read_engine (per-channel request)"),
            ("axi_rd_req_addr[63:0]","out","",           0, 0, "r_src_addr + offset (Q)",
             "to axi_read_engine"),
            ("axi_wr_req_valid","out","",                0, 0, "r_axi_wr_req (Q)",
             "to axi_write_engine"),
            ("axi_wr_req_addr[63:0]","out","",           0, 0, "r_dst_addr + offset (Q)",
             "to axi_write_engine"),
            ("ch_idle",        "out", "",                0, 1, "comb of r_state == IDLE",
             "to perf_profiler"),
            ("ch_done_strobe", "out", "",                0, 1, "r_state DONE-pulse (Q)",
             "to monbus_reporter"),
            ("monbus_pkt",     "out", "",                0, 1, "r_monbus_pkt event-encode (Q)",
             "to monbus group / reporter"),
        ]),
        ("axi_read_engine", [
            # local flops: r_arb_winner, r_active_ch, r_burst_id_ctr,
            # r_outstanding_per_ch[N], r_axi_ar_addr (pipe stage),
            # r_axi_ar_len, r_axi_ar_valid, r_axi_r_ready
            ("sched_req_valid[N-1:0]","in","",           1, 1, "u_arbiter (CLIENTS=N)",
             "from scheduler[N] (per-channel)"),
            ("sched_req_addr[N-1:0][63:0]","in","",      1, 0, "u_arbiter granted addr mux",
             "from scheduler[N]"),
            ("r_arb_winner[clog2(N)-1:0]","local","",    0, 1, "r_active_ch (D)",
             "round-robin output latch"),
            ("r_active_ch",    "local","",               0, 1, "r_axi_ar_addr / r_axi_ar_len (D)",
             "channel context for AR launch + R capture"),
            ("r_axi_ar_addr",  "local","",               0, 0, "u_rd_axi_skid.wr_data (D)",
             "pipelined ARADDR before skid"),
            ("r_axi_ar_valid", "local","",               0, 1, "u_rd_axi_skid.wr_valid (D)",
             "registered request to skid"),
            ("r_outstanding[N-1:0]","local","",          1, 1, "r_axi_ar_valid throttle",
             "per-channel in-flight counter"),
            ("r_burst_id_ctr", "local","",               0, 1, "r_axi_ar_addr id field (D)",
             "monotonic AXI ID generator"),
            ("m_axi_ar_*",     "out", "gaxi_skid_buffer",0, 0, "u_rd_axi_skid AR head (Q)",
             "to AXI4 master read system bus"),
            ("m_axi_r_*",      "in",  "gaxi_skid_buffer",0, 0, "u_rd_axi_skid R input",
             "from AXI4 master read system bus"),
            ("sram_alloc_req[N-1:0]","out","",           0, 0, "r_axi_ar_valid demux (Q)",
             "to sram_controller_unit[N] (reserve burst space)"),
            ("sram_alloc_space_free[N-1:0][AW:0]","in","",1, 0, "back-pressure cmp",
             "from sram_controller_unit[N]"),
            ("sram_wr_valid[N-1:0]","out","",            0, 0, "r_active_ch demux of R-skid valid (Q)",
             "to sram_controller_unit[N].wr"),
            ("sram_wr_data[DW-1:0]","out","",            0, 0, "u_rd_axi_skid.rd_data (Q)",
             "to sram_controller_unit[N].wr (channel-id muxed)"),
        ]),
        ("axi_write_engine", [
            # local flops: r_arb_winner, r_active_ch, r_axi_aw_addr,
            # r_axi_aw_len, r_axi_aw_valid, r_axi_w_data, r_axi_w_last,
            # r_axi_w_valid, r_b_phase_active, r_w_phase_active, r_drain_req
            ("sched_req_valid[N-1:0]","in","",           1, 1, "arbiter (D)",
             "from scheduler[N]"),
            ("sched_req_addr[N-1:0][63:0]","in","",      1, 0, "arbiter addr mux (D)",
             "from scheduler[N]"),
            ("sram_rd_valid[N-1:0]","in","",             1, 1, "r_w_phase_active (D)",
             "from sram_controller_unit[N] (data ready)"),
            ("sram_rd_data[N-1:0][DW-1:0]","in","",      1, 0, "r_axi_w_data (D)",
             "from sram_controller_unit[N]"),
            ("r_arb_winner",   "local","",               0, 1, "r_active_ch (D)",
             "round-robin output latch"),
            ("r_active_ch",    "local","",               0, 1, "r_axi_aw_addr / r_axi_w_data (D)",
             "channel context for AW launch + W stream"),
            ("r_axi_aw_addr",  "local","",               0, 0, "u_wr_axi_skid AW.wr_data (D)",
             "pipelined AWADDR"),
            ("r_axi_aw_valid", "local","",               0, 1, "u_wr_axi_skid AW.wr_valid (D)",
             "registered AW request"),
            ("r_axi_w_data",   "local","",               0, 0, "u_wr_axi_skid W.wr_data (D)",
             "W-beat data pipeline reg"),
            ("r_axi_w_last",   "local","",               0, 1, "u_wr_axi_skid W.wr_data WLAST bit (D)",
             "burst termination flag"),
            ("r_w_phase_active","local","",              0, 1, "r_drain_req (D)",
             "tracks burst in-flight; gates drain request"),
            ("r_b_phase_active","local","",              0, 1, "u_b_phase_txn_fifo wr enable (D)",
             "BRESP-await counter"),
            ("r_drain_req",    "local","",               0, 0, "sram_drain_req (Q)",
             "to sram_controller_unit"),
            ("sram_drain_req[N-1:0]","out","",           0, 0, "r_drain_req demux (Q)",
             "to sram_controller_unit[N]"),
            ("m_axi_aw_*",     "out", "gaxi_skid_buffer",0, 0, "u_wr_axi_skid AW head (Q)",
             "to AXI4 master write system bus"),
            ("m_axi_w_*",      "out", "gaxi_skid_buffer",0, 0, "u_wr_axi_skid W head (Q)",
             "to AXI4 master write system bus"),
            ("m_axi_b_*",      "in",  "gaxi_skid_buffer",0, 0, "u_wr_axi_skid B input",
             "from AXI4 master write system bus"),
            ("u_b_phase_txn_fifo wr","internal","gaxi_fifo_sync",0,0,"b-phase txn capture",
             "B-phase awaiting response"),
            ("u_w_phase_txn_fifo wr","internal","gaxi_fifo_sync",0,0,"w-phase txn capture",
             "W-phase awaiting last-beat"),
        ]),
    ]

    for fub_name, ports in FUBS:
        emit_fub_banner(fub_name)
        for (port, direction, bb, mux_lv, nand_lv, flop, goes) in ports:
            emit_row(fub_name, port, direction, bb, mux_lv, nand_lv, flop, goes)

    # Column widths
    widths = {1: 22, 2: 30, 3: 9,                    # FUB, port, clock
              4: 5, 5: 19, 6: 7, 7: 8,               # dir, BB, mux_lv, nand_lv
              8: 32, 9: 16, 10: 16, 11: 16,          # source flop, local delays
              12: 13,                                # delay_in
              13: 13, 14: 13, 15: 13,                # delay_out TT/FF/SS
              16: 3,                                 # blank gap
              17: 12, 18: 12, 19: 12,                # slack TT/FF/SS
              20: 50}                                # places it goes
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # Methodology notes at the bottom
    r = ws.max_row + 2
    ws.cell(row=r, column=1, value="Methodology").font = SECTION_FONT
    notes = [
        "One row per port (or significant internal handshake) per FUB. "
        "Hand-walked from the RTL.",
        "MUX lv / NAND lv = combinational gate-equivalents on the path within "
        "this FUB only - between the port and its nearest flop.",
        "When a port enters/exits a building block (gaxi_fifo_sync, "
        "gaxi_skid_buffer), the BB internal cost is captured on the 'building "
        "blocks' sheet - no double-counting here.",
        "'places it goes' lists the concrete consumer/producer in the wider "
        "STREAM hierarchy (stream_core / scheduler_group / APB CSR / system "
        "AXI bus).",
        "AXI bus signals (m_axi_ar_*, m_axi_r_*, m_axi_aw_*, m_axi_w_*, "
        "m_axi_b_*) are summarized as channel groups for brevity; each is a "
        "set of skid buffer outputs whose detailed cost is in the building "
        "blocks 'gaxi_skid_buffer' / 'axi4_master_rd/wr' rows.",
    ]
    for n in notes:
        r += 1
        ws.cell(row=r, column=1, value=n)
        ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=NCOL)


def main():
    probes = read_probes()
    per_level, tflop, single_pt = derive(probes)
    wb = Workbook()
    build_sheet(wb, per_level, tflop, single_pt)
    build_blocks_sheet(wb)
    build_stream_sheet(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  Tflop: {tflop}")
    print("  per-level (TT only):")
    for g in GATE_PRIMS:
        print(f"     {g:9s}: {per_level.get((g, 'TT'), 0):6.2f} ps")
    print("  degenerate single-point (TT, smallest size):")
    for g in DEGENERATE_PRIMS:
        d, lev = single_pt.get((g, "TT"), (0, 0))
        print(f"     {g:9s}: {d:6.2f} ps (lev={lev})")


if __name__ == "__main__":
    main()
