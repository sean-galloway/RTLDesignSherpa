#!/usr/bin/env python3
"""Build artix7_characterization.xlsx — the design-time prediction table for
the Xilinx Artix-7 family (Nexys A7-100T board uses xc7a100t).

Same 3-sheet shape as the ASIC companion (asap7_characterization.xlsx):

  characterization   per-primitive delay + Tflop + max-levels + wire/hop
                     + Design clocks table that downstream sheets pull from
  building blocks    per-block SWAG (skid / FIFO / arbiter / AXI master /
                     BRAM); clock-driven slack with bold-red conditional
                     formatting on negative

A third sheet 'measured (sweep)' is appended automatically when a
calibration CSV exists at work/sweep_wns.csv (produced by
`make calibrate` in fpga/).  It shows the measured WNS Vivado reported
at each target frequency so the designer can compare against the
predicted slack on the building-blocks sheet.

Seed data are Artix-7 primitive datasheet ballparks at -1/-2/-3 speed grades.
"""
from __future__ import annotations
import csv
import math
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

HERE = Path(__file__).resolve().parent
OUT  = HERE / "artix7_characterization.xlsx"
# Optional calibration CSV produced by `make calibrate` (or by hand).
# Same column shape as fpga/tools/parse_timing_sweep.py's output.
CALIBRATION_CSV = HERE / "sweep_wns.csv"

CORNERS = ["sg-1", "sg-2", "sg-3"]
FREQS   = [100, 150, 200, 250, 300]  # MHz - FPGA-typical sweep

PRIM_PS = {
    "LUT6":   {"sg-1": 200,  "sg-2": 170,  "sg-3": 140},
    "MUXF7":  {"sg-1": 250,  "sg-2": 220,  "sg-3": 180},
    "MUXF8":  {"sg-1": 280,  "sg-2": 250,  "sg-3": 210},
    "CARRY4": {"sg-1": 80,   "sg-2": 70,   "sg-3": 55},
    "DSP48":  {"sg-1": 2500, "sg-2": 2100, "sg-3": 1800},
    "BRAM":   {"sg-1": 2400, "sg-2": 2000, "sg-3": 1700},
    "LUTRAM": {"sg-1": 600,  "sg-2": 500,  "sg-3": 420},
    "SRL":    {"sg-1": 400,  "sg-2": 340,  "sg-3": 280},
}

PRIMS_IN_TABLE   = ["LUT6", "MUXF7", "MUXF8", "CARRY4", "DSP48", "LUTRAM", "SRL"]
DEGENERATE_PRIMS = ["BRAM"]
TFLOP_PROBE      = "LUT6"
MIXED_PROXY      = "DSP48"
TFLOP_PS         = {"sg-1": 400, "sg-2": 340, "sg-3": 280}

HDR_FONT       = Font(bold=True, color="000000")
HDR_FILL       = PatternFill("solid", fgColor="D9E1F2")
SECTION_FONT   = Font(bold=True)
SECTION_FILL   = PatternFill("solid", fgColor="D9E1F2")
PARAM_FILL     = PatternFill("solid", fgColor="FFF2CC")
FORMULA_FILL   = PatternFill("solid", fgColor="E7E6E6")
BIG_HDR_FONT   = Font(bold=True, size=14, color="000000")
BIG_HDR_FILL   = PatternFill("solid", fgColor="FFC000")
CENTER         = Alignment(horizontal="center")
THIN           = Side(border_style="thin", color="A0A0A0")
BOX            = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

CLOCK_TABLE_RANGE       = None
CLOCK_TABLE_FIRST_ROW_G = None
CLOCK_TABLE_LAST_ROW_G  = None


def section_header_row(ws, label, n_cols):
    ws.append([label] + [""] * (n_cols - 1))
    r = ws.max_row
    ws.cell(row=r, column=1).font = SECTION_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL


def build_characterization(wb):
    ws = wb.active
    ws.title = "characterization"

    col_headers = ["metric"] + [f"{c}_{f}MHz" for c in CORNERS for f in FREQS]
    ws.append(col_headers)
    for c in range(1, len(col_headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BOX

    row = ["Period (ps)"]
    for _ in CORNERS:
        for f in FREQS:
            row.append(round(1_000_000 / f, 1))
    ws.append(row)

    row = ["Clock-to-Out (Tcq + Tsu, ps)"]
    for sg in CORNERS:
        for _ in FREQS:
            row.append(TFLOP_PS[sg])
    ws.append(row)

    ws.append([""] * len(col_headers))
    section_header_row(ws, "Max primitive levels per cycle", len(col_headers))
    for prim in PRIMS_IN_TABLE:
        label = f"Max {prim} levels"
        if prim == MIXED_PROXY:
            label += "  (PROXY for mixed-bag logic)"
        row = [label]
        for sg in CORNERS:
            tf  = TFLOP_PS[sg]
            per = PRIM_PS[prim][sg]
            for f in FREQS:
                budget = (1_000_000 / f) - tf
                row.append(max(math.floor(budget / per), 0) if per > 0 else "")
        ws.append(row)
        if prim == MIXED_PROXY:
            for c in range(1, len(col_headers) + 1):
                ws.cell(row=ws.max_row, column=c).fill = PatternFill(
                    "solid", fgColor="FFF2CC")

    ws.append([""] * len(col_headers))
    section_header_row(ws, "Per-primitive delay (ps)  - freq-independent",
                       len(col_headers))
    for prim in PRIMS_IN_TABLE:
        label = f"per {prim} delay (ps)"
        if prim == MIXED_PROXY:
            label += "  (PROXY)"
        row = [label]
        for sg in CORNERS:
            for _ in FREQS:
                row.append(PRIM_PS[prim][sg])
        ws.append(row)
        if prim == MIXED_PROXY:
            for c in range(1, len(col_headers) + 1):
                ws.cell(row=ws.max_row, column=c).fill = PatternFill(
                    "solid", fgColor="FFF2CC")

    ws.append([""] * len(col_headers))
    section_header_row(
        ws,
        "Reference single-point primitives (ps)  -  not scaled by chain depth",
        len(col_headers))
    for prim in DEGENERATE_PRIMS:
        row = [f"{prim} access time (ps)"]
        for sg in CORNERS:
            for _ in FREQS:
                row.append(PRIM_PS[prim][sg])
        ws.append(row)

    ws.append([""] * len(col_headers))
    section_header_row(ws,
        "Design clocks  -  one row per clock; downstream sheets VLOOKUP by name",
        len(col_headers))
    r = ws.max_row + 1
    clock_hdr = ["clock name", "freq (MHz)", "target sg",
                 "period (ps)", "input delay 60% (ps)",
                 "output delay 40% (ps)", "wire/hop (ps)"]
    for i, h in enumerate(clock_hdr):
        cell = ws.cell(row=r, column=1 + i, value=h)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.alignment = CENTER

    global CLOCK_TABLE_RANGE, CLOCK_TABLE_FIRST_ROW_G, CLOCK_TABLE_LAST_ROW_G
    CLOCK_TABLE_FIRST_ROW_G = r + 1
    defaults = [
        ("clk_main", 100, "sg-1", 200),
        ("", "", "", ""),
        ("", "", "", ""),
        ("", "", "", ""),
        ("", "", "", ""),
        ("", "", "", ""),
    ]
    for i, (name, freq, sg, wire) in enumerate(defaults):
        rr = CLOCK_TABLE_FIRST_ROW_G + i
        c = ws.cell(row=rr, column=1, value=name);  c.fill = PARAM_FILL; c.alignment = CENTER
        c = ws.cell(row=rr, column=2, value=freq);  c.fill = PARAM_FILL; c.alignment = CENTER
        c = ws.cell(row=rr, column=3, value=sg);    c.fill = PARAM_FILL; c.alignment = CENTER
        c = ws.cell(row=rr, column=4,
                    value=f'=IF(ISNUMBER(B{rr}), 1000000/B{rr}, "")')
        c.fill = FORMULA_FILL; c.alignment = CENTER
        c = ws.cell(row=rr, column=5,
                    value=f'=IF(ISNUMBER(B{rr}), D{rr}*0.6, "")')
        c.fill = FORMULA_FILL; c.alignment = CENTER
        c = ws.cell(row=rr, column=6,
                    value=f'=IF(ISNUMBER(B{rr}), D{rr}*0.4, "")')
        c.fill = FORMULA_FILL; c.alignment = CENTER
        c = ws.cell(row=rr, column=7, value=wire); c.fill = PARAM_FILL; c.alignment = CENTER
    CLOCK_TABLE_LAST_ROW_G = CLOCK_TABLE_FIRST_ROW_G + len(defaults) - 1
    CLOCK_TABLE_RANGE = (
        f"characterization!$A${CLOCK_TABLE_FIRST_ROW_G}:$G${CLOCK_TABLE_LAST_ROW_G}"
    )

    ws.append([""] * len(col_headers))
    section_header_row(ws, "Notes", len(col_headers))
    notes = [
        "Speed grades: Artix-7 -1 (slowest, industrial) / -2 / -3 (fastest). "
        "Nexys A7-100T ships with xc7a100tcsg324-1.",
        "Tflop = FDRE Tcq + Tsu ~ 400 ps (-1).  Calibrate by reading actual "
        "FDRE timing from a real post-route timing_summary.txt.",
        "max_levels = floor((period - Tflop) / per_primitive_delay).",
        f"** When the path is a mixed bag of primitives, use the {MIXED_PROXY} "
        f"per-level number.  A DSP48 macro contains mux + adder + register, so "
        f"its per-level delay is a sensible blended estimate.",
        "BRAM is shown in the Reference section because its access time is a "
        "single-shot 1-cycle latency, not a scaled chain - use the absolute ps "
        "value, not levels-per-cycle.",
        "Wire delays land in the Design clocks table 'wire/hop' column; bump "
        "for longer routes or higher fanout, drop to 0 for paths that stay "
        "inside one CLB cluster.",
        "Seed data are Artix-7 datasheet + community ballparks.  Replace with "
        "calibrated numbers from real Vivado post-route reports (parse per-FUB "
        "WNS out of timing_summary.txt produced by `make bitstream`).",
    ]
    for n in notes:
        ws.append([n])

    ws.column_dimensions["A"].width = 40
    for c in range(2, len(col_headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12


def build_blocks_sheet(wb):
    ws = wb.create_sheet("building blocks")

    header = [
        "block",
        "P1", "P2", "P3", "P4", "P5", "P6",
        "params",
        "flop_count",
        "LUT_count_est",
        "in MUX lv", "in LUT lv",
        "data->D sg1 (ps)", "data->D sg2 (ps)", "data->D sg3 (ps)",
        "out MUX lv", "out LUT lv",
        "flop->out sg1 (ps)", "flop->out sg2 (ps)", "flop->out sg3 (ps)",
        "clock",
        "target period (ps)",
        "target sg",
        "target data->D (ps)",
        "target flop->out (ps)",
        "delay_in (ps)",
        "delay_out (ps)",
        "",
        "slack (ps)",
        "notes",
    ]
    ws.append(header)
    NCOL = len(header)
    for c in range(1, NCOL + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BOX

    # cols B..F = sg-1 (per_LUT=B15, per_MUX=B16);
    # cols G..K = sg-2; cols L..P = sg-3
    REFS = {
        "sg-1": {"lut": "$B$15", "mux": "$B$16"},
        "sg-2": {"lut": "$G$15", "mux": "$G$16"},
        "sg-3": {"lut": "$L$15", "mux": "$L$16"},
    }

    ws.cell(row=2, column=1, value="CONFIG").font = SECTION_FONT
    ws.cell(row=2, column=2, value=(
        "Each row's clock cell holds a name from characterization!Design "
        "clocks. Period / sg / wire/hop are VLOOKUP'd. Change the table "
        "to plan a new freq or speed grade and watch slack flip."))
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=NCOL)

    def emit_section_banner(label):
        r = ws.max_row + 1
        ws.append([label] + [""] * (NCOL - 1))
        ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=NCOL)
        cell = ws.cell(row=r, column=1)
        cell.font = BIG_HDR_FONT
        cell.fill = BIG_HDR_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 22

    SRAM_BLOCKS = [
        dict(name="BRAM (raw RAMB36)", params=[("W", 32), ("D", 1024)],
             flops=0, lut="=8*B{R} + 16*CEILING(LOG(C{R},2),1)",
             in_mux=0, in_lut=1, out_mux=0, out_lut=1,
             notes="RAMB36 macro (single-port, sync read). Storage W*D bits "
                   "lives in the macro, not LUTs. 1-cycle BRAM read latency."),
        dict(name="gaxi_fifo_sync (REG=1, BRAM-backed)",
             params=[("W", 32), ("D", 1024)],
             flops="=3*CEILING(LOG(C{R},2),1) + 1",
             lut="=B{R}*2 + CEILING(LOG(C{R},2),1)*4 + 16",
             in_mux=0, in_lut=1, out_mux=0, out_lut=1,
             notes="REGISTERED=1 + MEM_STYLE=BRAM: storage in RAMB36. "
                   "Used by SRAM controller units in STREAM/RAPIDS."),
    ]

    BUILDING_BLOCKS = [
        dict(name="gaxi_skid_buffer", params=[("W", 1), ("D", 2)],
             flops="=B{R}*C{R} + C{R} + 2",
             lut="=B{R}*C{R} + 8",
             in_mux=1, in_lut=1, out_mux=0, out_lut=0,
             notes="Skid head drives output directly (no out mux)."),
        dict(name="gaxi_fifo_sync (REG=0)", params=[("W", 1), ("D", 2)],
             flops="=B{R}*C{R} + 3*CEILING(LOG(C{R},2),1) + 1",
             lut="=B{R}*C{R}*2 + CEILING(LOG(C{R},2),1)*4 + 12",
             in_mux=1, in_lut=2,
             out_mux="=CEILING(LOG(C{R},2),1)", out_lut=0,
             notes="Output critical path = LUTRAM read mux tree, depth=log2(D)."),
        dict(name="arbiter_round_robin", params=[("N", 2)],
             flops="=B{R} + 2",
             lut="=B{R}*3",
             in_mux=0, in_lut=1,
             out_mux="=CEILING(LOG(B{R},2),1)", out_lut=1,
             notes="Priority encoder is log2(N) MUX levels on the output."),
        dict(name="axi4_master_rd",
             params=[("AW", 32), ("DW", 32), ("IW", 4), ("UW", 1),
                     ("SKID_AR", 2), ("SKID_R", 2)],
             flops="=F{R}*(B{R}+D{R}+E{R}+29) + G{R}*(D{R}+C{R}+E{R}+3)",
             lut="=2*(F{R}*(B{R}+D{R}+E{R}+29) + G{R}*(D{R}+C{R}+E{R}+3))",
             in_mux=1, in_lut=1, out_mux=1, out_lut=1,
             notes="Two skid buffers in series (AR + R)."),
        dict(name="axi4_master_wr",
             params=[("AW", 32), ("DW", 32), ("IW", 4), ("UW", 1),
                     ("SKID_AW", 2), ("SKID_W", 2)],
             flops="=F{R}*(B{R}+D{R}+E{R}+29) + G{R}*(C{R}+C{R}/8+E{R}+1) + 2*(D{R}+E{R}+2)",
             lut="=2*(F{R}*(B{R}+D{R}+E{R}+29) + G{R}*(C{R}+C{R}/8+E{R}+1) + 2*(D{R}+E{R}+2))",
             in_mux=1, in_lut=1, out_mux=1, out_lut=1,
             notes="Three skid buffers (AW + W + B); SKID_B fixed at 2."),
    ]

    def write_block(b):
        r = ws.max_row + 1
        params = b["params"]
        ws.cell(row=r, column=1, value=b["name"]).font = SECTION_FONT
        for i, (_lbl, v) in enumerate(params):
            cell = ws.cell(row=r, column=2 + i, value=v)
            cell.fill = PARAM_FILL; cell.alignment = CENTER
        ws.cell(row=r, column=8,
                value=", ".join(f"P{i+1}={lbl}" for i, (lbl, _) in enumerate(params)))
        flops_v = b["flops"]
        if isinstance(flops_v, str):
            flops_v = flops_v.format(R=r)
        ws.cell(row=r, column=9, value=flops_v)
        ws.cell(row=r, column=10, value=b["lut"].format(R=r))
        for col, key in ((11, "in_mux"), (12, "in_lut")):
            v = b[key]
            if isinstance(v, str): v = v.format(R=r)
            ws.cell(row=r, column=col, value=v)
        for off, sg in enumerate(("sg-1", "sg-2", "sg-3")):
            ref = REFS[sg]
            f = (f"=K{r}*characterization!{ref['mux']} "
                 f"+ L{r}*characterization!{ref['lut']}")
            c = ws.cell(row=r, column=13 + off, value=f)
            c.fill = FORMULA_FILL; c.alignment = CENTER
        for col, key in ((16, "out_mux"), (17, "out_lut")):
            v = b[key]
            if isinstance(v, str): v = v.format(R=r)
            ws.cell(row=r, column=col, value=v)
        for off, sg in enumerate(("sg-1", "sg-2", "sg-3")):
            ref = REFS[sg]
            f = (f"=P{r}*characterization!{ref['mux']} "
                 f"+ Q{r}*characterization!{ref['lut']}")
            c = ws.cell(row=r, column=18 + off, value=f)
            c.fill = FORMULA_FILL; c.alignment = CENTER
        c = ws.cell(row=r, column=21, value="clk_main")
        c.fill = PARAM_FILL; c.alignment = CENTER
        c = ws.cell(row=r, column=22,
                    value=f'=IFERROR(VLOOKUP(U{r}, {CLOCK_TABLE_RANGE}, 4, FALSE), "")')
        c.fill = FORMULA_FILL; c.alignment = CENTER
        c = ws.cell(row=r, column=23,
                    value=f'=IFERROR(VLOOKUP(U{r}, {CLOCK_TABLE_RANGE}, 3, FALSE), "")')
        c.fill = FORMULA_FILL; c.alignment = CENTER
        c = ws.cell(row=r, column=24,
                    value=(f'=IF(W{r}="sg-1",M{r},IF(W{r}="sg-2",N{r},'
                           f'IF(W{r}="sg-3",O{r},"")))'))
        c.fill = FORMULA_FILL; c.alignment = CENTER
        c = ws.cell(row=r, column=25,
                    value=(f'=IF(W{r}="sg-1",R{r},IF(W{r}="sg-2",S{r},'
                           f'IF(W{r}="sg-3",T{r},"")))'))
        c.fill = FORMULA_FILL; c.alignment = CENTER
        c = ws.cell(row=r, column=26,
                    value=f'=IFERROR(VLOOKUP(U{r}, {CLOCK_TABLE_RANGE}, 4, FALSE)*0.3, "")')
        c.fill = PARAM_FILL; c.alignment = CENTER
        c = ws.cell(row=r, column=27, value=f"=Z{r} + X{r} + Y{r}")
        c.fill = FORMULA_FILL; c.alignment = CENTER
        c = ws.cell(row=r, column=29,
                    value=(f'=IFERROR(V{r} - AA{r} - VLOOKUP(U{r}, '
                           f'{CLOCK_TABLE_RANGE}, 6, FALSE) - VLOOKUP(U{r}, '
                           f'{CLOCK_TABLE_RANGE}, 7, FALSE), "")'))
        c.fill = FORMULA_FILL; c.alignment = CENTER
        c.number_format = "0.000"
        ws.cell(row=r, column=30, value=b["notes"])

    emit_section_banner("BRAM-backed (SRAMs)")
    for b in SRAM_BLOCKS:
        write_block(b)
    emit_section_banner("Building Blocks")
    for b in BUILDING_BLOCKS:
        write_block(b)

    last_row = ws.max_row
    ws.conditional_formatting.add(
        f"AC3:AC{last_row}",
        CellIsRule(operator="lessThan", formula=["0"],
                   font=Font(bold=True, color="C00000")))

    widths = {1: 32, 8: 38, 9: 13, 10: 13,
              11: 8, 12: 9,
              13: 15, 14: 15, 15: 15,
              16: 8, 17: 9,
              18: 15, 19: 15, 20: 15,
              21: 11,
              22: 14, 23: 11, 24: 18, 25: 18,
              26: 13, 27: 13,
              28: 3,
              29: 13,
              30: 60}
    for c in range(2, 8):
        widths.setdefault(c, 9)
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


def build_measured_sheet(wb):
    """Append a 'measured (sweep)' sheet from work/sweep_wns.csv.

    Skips silently if no calibration CSV is present yet. The Makefile's
    `make calibrate` target writes one after a `bitstream-sweep` finishes.
    """
    if not CALIBRATION_CSV.exists():
        return
    rows = list(csv.DictReader(CALIBRATION_CSV.open()))
    if not rows:
        return
    ws = wb.create_sheet("measured (sweep)")
    cols = list(rows[0].keys())
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
    for r in rows:
        ws.append([r.get(c, "") for c in cols])
    # Negative WNS in red so failing sweep points jump out the same way
    # as the building-blocks slack column.
    last = ws.max_row
    wns_col = cols.index("wns_ns") + 1 if "wns_ns" in cols else None
    if wns_col:
        col_letter = get_column_letter(wns_col)
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{last}",
            CellIsRule(operator="lessThan", formula=["0"],
                       font=Font(bold=True, color="C00000")))
    for c in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18
    # Footer pointing back at the source.
    note_row = last + 2
    ws.cell(row=note_row, column=1, value=(
        f"Source: {CALIBRATION_CSV.name} (produced by `make calibrate` in "
        f"fpga/). Re-run that target after edits to refresh."))
    ws.cell(row=note_row, column=1).font = SECTION_FONT


def main():
    wb = Workbook()
    build_characterization(wb)
    build_blocks_sheet(wb)
    build_measured_sheet(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    sheets = ", ".join(wb.sheetnames)
    print(f"wrote {OUT}")
    print(f"  sheets: {sheets}")
    if not CALIBRATION_CSV.exists():
        print(f"  (no calibration CSV at {CALIBRATION_CSV.name} - skipped 'measured' sheet)")


if __name__ == "__main__":
    main()
