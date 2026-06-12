#!/usr/bin/env python3
r"""Generate budget_worksheet.xlsx -- the design-time budget sheet referenced
by the README's "From characterization to design" section.

ASIC-only.  Illustrative numbers throughout -- replace with real
characterization output once a target Liberty is selected.

Layout:
  README             -- how to use the workbook
  timings            -- per-structure timing + area card.  One row per
                        parameterised module / structure.  Columns split
                        out Tcq (clock-to-out), comb (internal
                        combinational delay), Tsu (setup), totals, flop
                        count, NAND-equivalent gate count, and per-period
                        slack at 3 ns / 4 ns / 5 ns.
  stream_write_eng   -- walk through the STREAM write-engine pipeline
                        using real module names.  Each row is one hop.
                        Some hops begin at a registered point (clean
                        clock edge); others consume a combinational
                        input from the previous hop and ride that path
                        further -- the "regd_in" column flags which is
                        which, and segment totals close out at the next
                        registered hop.  VLOOKUPs into the timings sheet
                        for delay / flops / gates.

The whole thing is a SWAG; the point is the shape of the methodology
and the chained-hop semantics, not the numbers.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

HERE = Path(__file__).resolve().parent
OUT  = HERE / "budget_worksheet.xlsx"

# ===========================================================================
# Per-structure timing + area card.
#
# Each row: (token, params, flops, nand_gates, Tcq_ps, comb_ps, Tsu_ps, notes)
# Tcq = clock-to-out delay at the structure's source flop.
# comb = combinational delay through the structure on the worst path.
# Tsu = setup time required at the structure's destination flop.
# total = Tcq + comb + Tsu  (the period the structure consumes if it
#         occupies an independent reg-to-reg hop).
# ===========================================================================
TIMINGS = [
    # ---- STREAM-write-engine modules (real names) -----------------------
    ("gaxi_skid_buffer(D=2, W=64)",
     "D=2 W=64",                 128,    640,   80,  120,  60,
     "skid; small AW/cmd payload"),
    ("gaxi_skid_buffer(D=2, W=512)",
     "D=2 W=512",                1024,   5120,  80,  130,  60,
     "skid; full W datapath"),
    ("gaxi_skid_buffer(D=2, W=8)",
     "D=2 W=8",                  16,     128,   80,  110,  60,
     "skid; B response payload"),
    ("gaxi_fifo_sync(D=8, W=64)",
     "D=8 W=64",                 512,    3200,  90,  220,  70,
     "descriptor cmd fifo"),
    ("gaxi_fifo_sync(D=4, W=8)",
     "D=4 W=8",                  32,     420,   90,  180,  70,
     "B response fifo"),
    ("arbiter_round_robin(N=8)",
     "N=8",                      24,     96,    70,  140,  60,
     "channel RR arbiter"),
    ("axi_gen_addr(AW=64)",
     "AW=64",                    0,      280,   0,   480,  60,
     "comb burst-addr gen (no internal flop)"),
    ("sram_read_mux(N=8, W=512)",
     "N=8 W=512",                512,    4096,  70,  290,  60,
     "8:1 data mux at SRAM read"),
    ("sram_drain_ctrl(NC=8)",
     "NC=8",                     96,     360,   70,  220,  60,
     "per-channel drain bookkeeping"),
    ("descriptor_decode",
     "(comb)",                   0,      210,   0,   180,  60,
     "comb decode of next descriptor"),
    ("write_engine_fsm",
     "S=4",                      18,     120,   70,  150,  60,
     "tiny FSM, 4 states"),
    # ---- Generic primitives (for reuse across paths) --------------------
    ("flop(W=64)",
     "W=64",                     64,     320,   60,  0,    50,
     "register (Tcq + Tsu only)"),
    ("flop(W=512)",
     "W=512",                    512,    2560,  60,  0,    50,
     "register (Tcq + Tsu only)"),
    ("mux(8:1, W=64)",
     "M=8 W=64",                 0,      448,   0,   150,  50,
     "generic comb mux"),
    ("adder(W=32)",
     "W=32",                     0,      96,    0,   485,  50,
     "ripple adder"),
]

# Convenience: column widths for the timings sheet.
TIMINGS_COLS = [
    ("structure",      30),
    ("params",         16),
    ("flops",          8),
    ("nand_gates",     11),
    ("Tcq_ps",         8),
    ("comb_ps",        9),
    ("Tsu_ps",         8),
    ("total_ps",       10),
    ("slack @3ns ps",  14),
    ("slack @4ns ps",  14),
    ("slack @5ns ps",  14),
    ("notes",          40),
]


# ===========================================================================
# STREAM write-engine pipeline walk.
#
# Each row is one hop.  Columns:
#   step       : pipeline label / annotation
#   module     : structure token -- must match a TIMINGS row
#   regd_in    : True if this hop's input arrives off a flop (begins a
#                new register-to-register segment); False if it consumes
#                a combinational input from the previous hop
#   notes      : anything worth saying
#
# Convention: when regd_in=True, the segment "starts" at this row's
# source flop (Tcq).  Subsequent rows with regd_in=False ride the same
# combinational segment; their comb_ps accumulates and their Tcq is
# ignored.  The segment "closes" at the NEXT regd_in=True row (or end
# of path), where that row's Tsu sets up the destination flop.
# ===========================================================================
WRITE_ENGINE_PATHS = [
    # Path A: AXI write command path  (scheduler descriptor -> m_axi_aw)
    ("Path A: descriptor -> m_axi_aw (command)", [
        ("desc reg out of scheduler",       "flop(W=64)",
         True,   "descriptor latched in scheduler FSM"),
        ("descriptor_decode",               "descriptor_decode",
         False,  "comb decode (NOT registered)"),
        ("descriptor cmd FIFO push",        "gaxi_fifo_sync(D=8, W=64)",
         True,   "FIFO breaks the segment"),
        ("RR channel select",               "arbiter_round_robin(N=8)",
         False,  "comb arbiter output drives addr gen"),
        ("burst addr gen",                  "axi_gen_addr(AW=64)",
         False,  "comb addr math (no flop)"),
        ("AW skid push",                    "gaxi_skid_buffer(D=2, W=64)",
         True,   "skid output drives m_axi_aw"),
    ]),
    # Path B: AXI write data path  (sram_controller -> m_axi_w)
    ("Path B: SRAM read -> m_axi_w (data)", [
        ("sram_drain_ctrl out",             "sram_drain_ctrl(NC=8)",
         True,   "registered drain ID"),
        ("per-channel data mux",            "sram_read_mux(N=8, W=512)",
         False,  "comb 8:1 mux into the W stream"),
        ("W skid push",                     "gaxi_skid_buffer(D=2, W=512)",
         True,   "skid output drives m_axi_w"),
    ]),
    # Path C: B-response collection  (m_axi_b -> done queue)
    ("Path C: m_axi_b -> b_done_q", [
        ("B skid out",                      "gaxi_skid_buffer(D=2, W=8)",
         True,   "skid output of m_axi_b"),
        ("write_engine FSM",                "write_engine_fsm",
         False,  "comb decode of B response"),
        ("B response FIFO push",            "gaxi_fifo_sync(D=4, W=8)",
         True,   "B done event queued"),
    ]),
]


# ===========================================================================
# Styling
# ===========================================================================
HDR_FILL    = PatternFill("solid", fgColor="DDDDDD")
PATH_FILL   = PatternFill("solid", fgColor="EAF0F8")
SEG_FILL    = PatternFill("solid", fgColor="F8F8F8")
TOTAL_FILL  = PatternFill("solid", fgColor="F2D7B4")
SLACK_OK    = PatternFill("solid", fgColor="C8E6C9")
SLACK_BAD   = PatternFill("solid", fgColor="FFC0C0")
BOLD        = Font(bold=True)
ITALIC      = Font(italic=True, color="666666")
MONO        = Font(name="Consolas")
THIN        = Side(border_style="thin", color="888888")
BOX         = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ===========================================================================
# Sheet 1: README
# ===========================================================================
def build_readme(ws):
    ws.append(["STREAM write-engine illustrative budget"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    for line in [
        "Purpose",
        "  Walk through the STREAM write engine end-to-end using real",
        "  module names, real-ish flop / NAND-gate counts, and SWAG",
        "  timing numbers, so the designer can SEE that the methodology",
        "  catches surprises BEFORE synthesis.",
        "",
        "Sheets",
        "  1. timings           -- per-structure card (Tcq, comb, Tsu,",
        "                          flops, nand_gates), one row per",
        "                          parameterised instance, plus slack",
        "                          at periods 3 / 4 / 5 ns.",
        "  2. stream_write_eng  -- 3 paths through the write engine.",
        "                          Each row = one hop.  regd_in flags",
        "                          whether the hop starts a new",
        "                          register-to-register segment.  At",
        "                          each segment close, total = Tcq +",
        "                          sum(comb) + Tsu and slack is the",
        "                          period minus total.",
        "",
        "Why segments?",
        "  Mid-pipeline FUBs often consume a combinational input from",
        "  the previous stage (no intervening flop).  The path budget",
        "  must chain those FUBs onto the same Tcq -> Tsu segment,",
        "  not treat each one as an independent hop.  When regd_in is",
        "  False, this worksheet keeps the running comb total alive;",
        "  when regd_in is True, the previous segment closes and a",
        "  new one starts at the new source flop.",
        "",
        "Editing rules",
        "  - Change the 'Period (ps)' cell at the top of",
        "    stream_write_eng to sweep frequency.  Slack cells",
        "    recompute and the conditional formatting flips green/red.",
        "  - Add a hop: insert a row, fill in module token (must match",
        "    a row in 'timings'), regd_in flag, and notes.",
        "  - Add a structure: extend the TIMINGS table.  The path",
        "    worksheet's VLOOKUPs pick it up on next recalc.",
        "",
        "Caveats",
        "  - All numbers are illustrative SWAGs (5-nm-ish stdcells",
        "    assumed).  Replace with real characterization output.",
        "  - FPGA characterization is the subject of a separate addendum",
        "    README; the methodology is the same but the numbers differ.",
    ]:
        ws.append([line])
    ws.column_dimensions["A"].width = 78


# ===========================================================================
# Sheet 2: timings
# ===========================================================================
def build_timings(ws):
    # Row 1: column headers
    for i, (name, width) in enumerate(TIMINGS_COLS, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.font, c.fill, c.alignment, c.border = BOLD, HDR_FILL, Alignment(horizontal="center"), BOX
        ws.column_dimensions[get_column_letter(i)].width = width

    # Rows: one per structure
    for r, row in enumerate(TIMINGS, start=2):
        token, params, flops, nand, tcq, comb, tsu, notes = row
        ws.cell(row=r, column=1, value=token).font = MONO
        ws.cell(row=r, column=2, value=params).font = ITALIC
        ws.cell(row=r, column=3, value=flops)
        ws.cell(row=r, column=4, value=nand)
        ws.cell(row=r, column=5, value=tcq)
        ws.cell(row=r, column=6, value=comb)
        ws.cell(row=r, column=7, value=tsu)
        # total = Tcq + comb + Tsu
        ws.cell(row=r, column=8, value=f"=E{r}+F{r}+G{r}").font = BOLD
        # Per-period slack
        ws.cell(row=r, column=9,  value=f"=3000-H{r}")
        ws.cell(row=r, column=10, value=f"=4000-H{r}")
        ws.cell(row=r, column=11, value=f"=5000-H{r}")
        ws.cell(row=r, column=12, value=notes).font = ITALIC
        for c in range(1, 13):
            ws.cell(row=r, column=c).border = BOX

    # Conditional formatting on slack columns
    from openpyxl.formatting.rule import CellIsRule
    last = len(TIMINGS) + 1
    for col_letter in ("I", "J", "K"):
        rng = f"{col_letter}2:{col_letter}{last}"
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="lessThan", formula=["0"], fill=SLACK_BAD))
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=SLACK_OK))

    ws.freeze_panes = "A2"


# ===========================================================================
# Sheet 3: stream_write_eng
# ===========================================================================
def build_path_worksheet(ws, timings_last_row: int):
    # ---- Configuration block ------------------------------------------
    ws["A1"] = "Configuration"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "Period (ps)"
    ws["B2"] = 4000  # default to 4 ns / 250 MHz
    ws["A3"] = "Target freq (MHz)"
    ws["B3"] = "=1e6/B2"
    for r in (2, 3):
        ws[f"A{r}"].font, ws[f"A{r}"].fill = BOLD, HDR_FILL

    # ---- Header row ----------------------------------------------------
    HEADERS = [
        "step", "module token", "regd_in", "Tcq_ps", "comb_ps", "Tsu_ps",
        "seg_total_ps", "seg_slack_ps", "flops", "nand_gates", "notes",
    ]
    HDR_ROW = 5
    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=HDR_ROW, column=i, value=h)
        c.font, c.fill, c.alignment, c.border = BOLD, HDR_FILL, Alignment(horizontal="center"), BOX
    ws.column_dimensions["A"].width = 32   # step
    ws.column_dimensions["B"].width = 30   # module
    ws.column_dimensions["C"].width = 8    # regd_in
    for col in ("D", "E", "F", "G", "H"):
        ws.column_dimensions[col].width = 11
    ws.column_dimensions["I"].width = 7    # flops
    ws.column_dimensions["J"].width = 11   # gates
    ws.column_dimensions["K"].width = 50   # notes

    tim_range = f"timings!$A$2:$L${timings_last_row}"
    period_cell = "$B$2"

    row = HDR_ROW + 1
    total_flops_cells = []
    total_gate_cells  = []
    seg_slack_cells   = []

    for path_label, hops in WRITE_ENGINE_PATHS:
        # Path label row
        ws.cell(row=row, column=1, value=path_label).font = Font(bold=True, italic=True)
        ws.cell(row=row, column=1).fill = PATH_FILL
        for c in range(1, len(HEADERS) + 1):
            ws.cell(row=row, column=c).border = BOX
        row += 1

        # Hop rows (data only; segment totals are emitted as separate
        # rows when a segment closes).
        seg_start_row = None      # row where current segment began
        for hop_idx, hop in enumerate(hops):
            step, token, regd_in, notes = hop

            # If this hop starts a new segment and a previous segment is
            # open, close the previous segment FIRST (sum its comb_ps,
            # add prev Tcq + this Tsu, slack vs period).
            # We handle segment close at the row AFTER the last hop in
            # the segment -- easiest to emit the close-row inline.
            if regd_in and seg_start_row is not None:
                _emit_segment_close(ws, seg_start_row, row - 1, period_cell,
                                    seg_slack_cells)
                row += 1   # blank-ish for the closing row
                seg_start_row = None

            if regd_in:
                seg_start_row = row

            # Module token + lookups
            ws.cell(row=row, column=1, value=step)
            ws.cell(row=row, column=2, value=token).font = MONO
            ws.cell(row=row, column=3,
                    value="yes" if regd_in else "no").alignment = Alignment(horizontal="center")
            # Tcq / comb / Tsu via VLOOKUP into timings (cols 5/6/7)
            ws.cell(row=row, column=4,
                    value=f'=IFERROR(VLOOKUP(B{row},{tim_range},5,FALSE),0)')
            ws.cell(row=row, column=5,
                    value=f'=IFERROR(VLOOKUP(B{row},{tim_range},6,FALSE),0)')
            ws.cell(row=row, column=6,
                    value=f'=IFERROR(VLOOKUP(B{row},{tim_range},7,FALSE),0)')
            # flops / nand from timings (cols 3/4)
            ws.cell(row=row, column=9,
                    value=f'=IFERROR(VLOOKUP(B{row},{tim_range},3,FALSE),0)')
            ws.cell(row=row, column=10,
                    value=f'=IFERROR(VLOOKUP(B{row},{tim_range},4,FALSE),0)')
            ws.cell(row=row, column=11, value=notes).font = ITALIC
            for c in range(1, len(HEADERS) + 1):
                ws.cell(row=row, column=c).border = BOX

            total_flops_cells.append(f"I{row}")
            total_gate_cells.append(f"J{row}")

            row += 1

        # Close the final segment of this path
        if seg_start_row is not None:
            _emit_segment_close(ws, seg_start_row, row - 1, period_cell,
                                seg_slack_cells)
            row += 1

        # Blank row between paths
        row += 1

    # ---- Grand totals --------------------------------------------------
    ws.cell(row=row, column=1, value="GRAND TOTALS").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = TOTAL_FILL
    if seg_slack_cells:
        ws.cell(row=row, column=8, value=f"=MIN({','.join(seg_slack_cells)})").font = BOLD
        ws.cell(row=row, column=8).fill = TOTAL_FILL
    ws.cell(row=row, column=9, value=f"=SUM({','.join(total_flops_cells)})").font = BOLD
    ws.cell(row=row, column=9).fill = TOTAL_FILL
    ws.cell(row=row, column=10, value=f"=SUM({','.join(total_gate_cells)})").font = BOLD
    ws.cell(row=row, column=10).fill = TOTAL_FILL

    # ---- Conditional formatting on the slack column ----
    from openpyxl.formatting.rule import CellIsRule
    slack_rng = f"H{HDR_ROW + 1}:H{row}"
    ws.conditional_formatting.add(
        slack_rng, CellIsRule(operator="lessThan", formula=["0"], fill=SLACK_BAD))
    ws.conditional_formatting.add(
        slack_rng, CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=SLACK_OK))

    ws.freeze_panes = "C6"


def _emit_segment_close(ws, seg_start_row, seg_end_row, period_cell, slack_cells):
    """Emit a single 'segment total' row immediately after seg_end_row.
    Total = Tcq(start) + SUM(comb from start..end) + Tsu(end)
    Slack = period_cell - total"""
    close_row = seg_end_row + 1
    ws.cell(row=close_row, column=1, value="    segment total").font = Font(italic=True, bold=True)
    ws.cell(row=close_row, column=1).fill = SEG_FILL
    total = (f"=D{seg_start_row}"
             f"+SUM(E{seg_start_row}:E{seg_end_row})"
             f"+F{seg_end_row}")
    ws.cell(row=close_row, column=7, value=total).font = BOLD
    ws.cell(row=close_row, column=7).fill = TOTAL_FILL
    slack = f"={period_cell}-G{close_row}"
    ws.cell(row=close_row, column=8, value=slack).font = BOLD
    ws.cell(row=close_row, column=8).fill = TOTAL_FILL
    for c in range(1, 12):
        ws.cell(row=close_row, column=c).border = BOX
    slack_cells.append(f"H{close_row}")


# ===========================================================================
# Driver
# ===========================================================================
def main() -> None:
    wb = Workbook()

    readme = wb.active
    readme.title = "README"
    build_readme(readme)

    timings = wb.create_sheet("timings")
    build_timings(timings)
    timings_last_row = len(TIMINGS) + 1  # data rows + header

    path = wb.create_sheet("stream_write_eng")
    build_path_worksheet(path, timings_last_row)

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
